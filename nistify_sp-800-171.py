#!/usr/bin/env python3
"""
NISTify 800-171 - NIST SP 800-171 Compliance Scanner and Reporter
Windows Compatible Version (No Emojis)
Scans networks and endpoints for compliance with NIST SP 800-171 Revision 2 and Revision 3
Generates compliance reports in multiple formats and POA&M documents
"""

import sys
import json
import subprocess
import platform
import datetime
import argparse
import logging
from pathlib import Path
from dataclasses import dataclass, asdict
from typing import List, Dict, Optional, Tuple
import ipaddress

# Third-party imports
try:
    import nmap
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    import networkx as nx
    import matplotlib.pyplot as plt
    import matplotlib.patches as mpatches 
except ImportError as e:
    print(f"Missing required package: {e}")
    print("Install with: pip install python-nmap openpyxl networkx matplotlib")
    sys.exit(1)

def print_banner():
    """Print ASCII art banner for NISTify 800-171 R2"""
    banner = """
    ╔══════════════════════════════════════════════════════════════════════════════════════════╗
    ║                                                                                          ║
    ║    ███╗   ██╗██╗███████╗████████╗██╗███████╗██╗   ██╗                                    ║
    ║    ████╗  ██║██║██╔════╝╚══██╔══╝██║██╔════╝╚██╗ ██╔╝                                    ║
    ║    ██╔██╗ ██║██║███████╗   ██║   ██║█████╗   ╚████╔╝                                     ║
    ║    ██║╚██╗██║██║╚════██║   ██║   ██║██╔══╝    ╚██╔╝                                      ║
    ║    ██║ ╚████║██║███████║   ██║   ██║██║        ██║                                       ║
    ║    ╚═╝  ╚═══╝╚═╝╚══════╝   ╚═╝   ╚═╝╚═╝        ╚═╝                                       ║
    ║                                                          By: Nightstalker                ║
    ║              ╔══════════════════════════════════════════════════╗                        ║
    ║              ║         800-171 Rev 2  |  Rev 3                  ║                        ║
    ║              ╚══════════════════════════════════════════════════╝                        ║
    ║                                                                                          ║
    ║        NIST SP 800-171 Rev 2 / Rev 3 Compliance Scanner & Assessment Tool                ║
    ║                                                                                          ║
    ║   ┌─────────────────────────────────────────────────────────────────────────────────┐    ║
    ║   │  * Automated Network Discovery & Port Scanning                                  │    ║
    ║   │  * NIST SP 800-171 Rev 2 and Rev 3 Compliance Assessment                        │    ║
    ║   │  * SPRS Score Calculation & Risk Analysis                                       │    ║
    ║   │  * Network Topology Visualization                                               │    ║
    ║   │  * Multi-Format Reporting (HTML, JSON, Excel, Text, Nmap)                       │    ║
    ║   │  * Plan of Action & Milestones (POA&M) Generation                               │    ║
    ║   └─────────────────────────────────────────────────────────────────────────────────┘    ║
    ║                                                                                          ║
    ║               Version: 2.0.0  |  License: GPL-3.0  |  Windows Compatible                 ║
    ║                                                                                          ║
    ╚══════════════════════════════════════════════════════════════════════════════════════════╝

    """
    print(banner)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('nistify800-171r2.log', encoding='utf-8'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Verbose screen-output helpers
# ---------------------------------------------------------------------------
_verbose_mode: bool = False

_SEV_LABEL = {
    "high":   "[HIGH]  ",
    "medium": "[MED]   ",
    "low":    "[LOW]   ",
}
_STATUS_LABEL = {
    "compliant":              "[PASS]",
    "non_compliant":          "[FAIL]",
    "MANUAL_REVIEW_REQUIRED": "[MANUAL REVIEW]",
    "not_assessed":           "[NOT ASSESSED]",
    "not_applicable":         "[N/A]",
}

def _vprint(msg: str = "", indent: int = 0) -> None:
    """Print a line to stdout only when verbose mode is active."""
    if _verbose_mode:
        print("  " * indent + msg)

def _vsection(title: str) -> None:
    """Print a bold section header (80-char bar) in verbose mode."""
    if not _verbose_mode:
        return
    bar = "=" * 80
    print(f"\n{bar}")
    print(f"  {title}")
    print(bar)

def _vsubsection(title: str) -> None:
    """Print a lighter sub-section divider in verbose mode."""
    if not _verbose_mode:
        return
    bar = "-" * 80
    print(f"\n{bar}")
    print(f"  {title}")
    print(bar)

def _vkv(label: str, value, indent: int = 1, width: int = 20) -> None:
    """Print a label/value pair in verbose mode."""
    if _verbose_mode:
        pad = "  " * indent
        print(f"{pad}{label:<{width}}: {value}")

# ---------------------------------------------------------------------------

@dataclass
class ComplianceResult:
    """Result of a compliance check"""
    control_id: str
    control_name: str
    control_text: str
    status: str
    finding: str
    remediation: str
    severity: str
    evidence: List[str]

@dataclass
class NetworkTopology:
    """Network topology information"""
    nodes: Dict[str, Dict]
    edges: List[Tuple[str, str]]
    subnets: List[str]
    gateways: List[str]
    network_diagram_path: Optional[str] = None

@dataclass
class SystemInfo:
    """System information for scanned endpoint"""
    hostname: str
    ip_address: str
    os_type: str
    os_version: str
    open_ports: List[int]
    services: Dict[int, str]
    last_scanned: str
    mac_address: Optional[str] = None
    vendor: Optional[str] = None
    hop_count: Optional[int] = None
    gateway: Optional[str] = None

class NIST80171Controls:
    """NIST SP 800-171 Rev 2 Control Definitions"""

    CONTROLS = {
    "3.1.1": {
    "name": "Access Control Policy and Procedures",
    "text": "Limit system access to authorized users, processes acting on behalf of authorized users, and devices (including other systems).",
    "family": "AC"
    },
    "3.1.2": {
        "name": "Transaction and Function Control",
        "text": "Limit system access to the types of transactions and functions that authorized users are permitted to execute.",
        "family": "AC"
    },
    "3.1.3": {
        "name": "External Connections",
        "text": "Control the flow of CUI in accordance with approved authorizations.",
        "family": "AC"
    },
    "3.1.4": {
        "name": "Separation of Duties",
        "text": "Separate the duties of individuals to reduce the risk of malevolent activity without collusion.",
        "family": "AC"
    },
    "3.1.5": {
        "name": "Least Privilege",
        "text": "Employ the principle of least privilege, including for specific security functions and privileged accounts.",
        "family": "AC"
    },
    "3.1.6": {
        "name": "Non-Privileged Account Use",
        "text": "Use non-privileged accounts or roles when accessing nonsecurity functions.",
        "family": "AC"
    },
    "3.1.7": {
        "name": "Privileged Functions",
        "text": "Prevent non-privileged users from executing privileged functions and capture the execution of such functions in audit logs.",
        "family": "AC"
    },
    "3.1.8": {
        "name": "Unsuccessful Logon Attempts",
        "text": "Limit unsuccessful logon attempts.",
        "family": "AC"
    },
    "3.1.9": {
        "name": "Privacy and Security Notices",
        "text": "Provide privacy and security notices consistent with applicable CUI rules.",
        "family": "AC"
    },
    "3.1.10": {
        "name": "Session Lock",
        "text": "Use session lock with pattern-hiding displays to prevent access and viewing of data after a period of inactivity.",
        "family": "AC"
    },
    "3.1.11": {
        "name": "Session Termination",
        "text": "Terminate (automatically) a user session after a defined condition.",
        "family": "AC"
    },
    "3.1.12": {
        "name": "Control Remote Access",
        "text": "Monitor and control remote access sessions.",
        "family": "AC"
    },
    "3.1.13": {
        "name": "Remote Access Confidentiality",
        "text": "Employ cryptographic mechanisms to protect the confidentiality of remote access sessions.",
        "family": "AC"
    },
    "3.1.14": {
        "name": "Remote Access Routing",
        "text": "Route remote access via managed access control points.",
        "family": "AC"
    },
    "3.1.15": {
        "name": "Privileged Remote Access",
        "text": "Authorize remote execution of privileged commands and remote access to security-relevant information.",
        "family": "AC"
    },
    "3.1.16": {
        "name": "Wireless Access Authorization",
        "text": "Authorize wireless access prior to allowing such connections.",
        "family": "AC"
    },
    "3.1.17": {
        "name": "Wireless Access Protection",
        "text": "Protect wireless access using authentication and encryption.",
        "family": "AC"
    },
    "3.1.18": {
        "name": "Mobile Device Connection",
        "text": "Control connection of mobile devices.",
        "family": "AC"
    },
    "3.1.19": {
        "name": "Encrypt CUI on Mobile",
        "text": "Encrypt CUI on mobile devices and mobile computing platforms.",
        "family": "AC"
    },
    "3.1.20": {
        "name": "External System Use",
        "text": "Verify and control/limit connections to and use of external systems.",
        "family": "AC"
    },
    "3.1.21": {
        "name": "Portable Storage Device Use",
        "text": "Limit use of portable storage devices on external systems.",
        "family": "AC"
    },
    "3.1.22": {
        "name": "Publicly Accessible System Content",
        "text": "Control CUI posted or processed on publicly accessible systems.",
        "family": "AC"
    },

    # 3.2 AWARENESS AND TRAINING (AT)
    "3.2.1": {
        "name": "Security Awareness",
        "text": "Ensure that managers, systems administrators, and users of organizational systems are made aware of the security risks associated with their activities and of the applicable policies, standards, and procedures related to the security of those systems.",
        "family": "AT"
    },
    "3.2.2": {
        "name": "Insider Threat Awareness",
        "text": "Ensure that personnel are trained to carry out their assigned information security-related duties and responsibilities.",
        "family": "AT"
    },
    "3.2.3": {
        "name": "Security Training",
        "text": "Provide security awareness training on recognizing and reporting potential indicators of insider threat.",
        "family": "AT"
    },

    # 3.3 AUDIT AND ACCOUNTABILITY (AU)
    "3.3.1": {
        "name": "System Audit Records",
        "text": "Create and retain system audit logs and records to the extent needed to enable the monitoring, analysis, investigation, and reporting of unlawful or unauthorized system activity.",
        "family": "AU"
    },
    "3.3.2": {
        "name": "Audit Record Content",
        "text": "Ensure that the actions of individual system users can be uniquely traced to those users so they can be held accountable for their actions.",
        "family": "AU"
    },
    "3.3.3": {
        "name": "Audit Record Review",
        "text": "Review and update logged events.",
        "family": "AU"
    },
    "3.3.4": {
        "name": "Audit Failure Response",
        "text": "Alert in the event of an audit logging process failure.",
        "family": "AU"
    },
    "3.3.5": {
        "name": "Audit Correlation",
        "text": "Correlate audit record review, analysis, and reporting processes for investigation and response to indications of unlawful, unauthorized, suspicious, or unusual activity.",
        "family": "AU"
    },
    "3.3.6": {
        "name": "Audit Reduction",
        "text": "Provide audit record reduction and report generation to support on-demand analysis and reporting.",
        "family": "AU"
    },
    "3.3.7": {
        "name": "Audit Monitoring",
        "text": "Provide a system capability that compares and synchronizes internal system clocks with an authoritative source to generate time stamps for audit records.",
        "family": "AU"
    },
    "3.3.8": {
        "name": "Audit Record Protection",
        "text": "Protect audit information and audit logging tools from unauthorized access, modification, and deletion.",
        "family": "AU"
    },
    "3.3.9": {
        "name": "Audit Management",
        "text": "Limit management of audit logging functionality to a subset of privileged users.",
        "family": "AU"
    },

    # 3.4 CONFIGURATION MANAGEMENT (CM)
    "3.4.1": {
        "name": "Baseline Configuration",
        "text": "Establish and maintain baseline configurations and inventories of organizational systems (including hardware, software, firmware, and documentation) throughout the respective system development life cycles.",
        "family": "CM"
    },
    "3.4.2": {
        "name": "Security Configuration Settings",
        "text": "Establish and enforce security configuration settings for information technology products employed in organizational systems.",
        "family": "CM"
    },
    "3.4.3": {
        "name": "Configuration Change Control",
        "text": "Track, review, approve or disapprove, and log changes to organizational systems.",
        "family": "CM"
    },
    "3.4.4": {
        "name": "Security Impact Analysis",
        "text": "Analyze the security impact of changes prior to implementation.",
        "family": "CM"
    },
    "3.4.5": {
        "name": "Access Restrictions",
        "text": "Define, document, approve, and enforce physical and logical access restrictions associated with changes to organizational systems.",
        "family": "CM"
    },
    "3.4.6": {
        "name": "Least Functionality",
        "text": "Employ the principle of least functionality by configuring organizational systems to provide only essential capabilities.",
        "family": "CM"
    },
    "3.4.7": {
        "name": "Nonessential Functionality",
        "text": "Restrict, disable, or prevent the use of nonessential programs, functions, ports, protocols, and services.",
        "family": "CM"
    },
    "3.4.8": {
        "name": "Application Execution Policy",
        "text": "Apply deny-by-exception (blacklisting) policy to prevent the use of unauthorized software or deny-all, permit-by-exception (whitelisting) policy to allow the execution of authorized software.",
        "family": "CM"
    },
    "3.4.9": {
        "name": "User-Installed Software",
        "text": "Control and monitor user-installed software.",
        "family": "CM"
    },

    # 3.5 IDENTIFICATION AND AUTHENTICATION (IA)
    "3.5.1": {
        "name": "User Identification",
        "text": "Identify system users, processes acting on behalf of users, and devices.",
        "family": "IA"
    },
    "3.5.2": {
        "name": "User Authentication",
        "text": "Authenticate (or verify) the identities of users, processes, or devices, as a prerequisite to allowing access to organizational systems.",
        "family": "IA"
    },
    "3.5.3": {
        "name": "Multifactor Authentication",
        "text": "Use multifactor authentication for local and network access to privileged accounts and for network access to non-privileged accounts.",
        "family": "IA"
    },
    "3.5.4": {
        "name": "Replay-Resistant Authentication",
        "text": "Employ replay-resistant authentication mechanisms for network access to privileged and non-privileged accounts.",
        "family": "IA"
    },
    "3.5.5": {
        "name": "Identifier Management",
        "text": "Prevent reuse of identifiers for a defined period.",
        "family": "IA"
    },
    "3.5.6": {
        "name": "Authenticator Management",
        "text": "Disable identifiers after a defined period of inactivity.",
        "family": "IA"
    },
    "3.5.7": {
        "name": "Password Complexity",
        "text": "Enforce a minimum password complexity and change of characters when new passwords are created.",
        "family": "IA"
    },
    "3.5.8": {
        "name": "Password Reuse",
        "text": "Prohibit password reuse for a specified number of generations.",
        "family": "IA"
    },
    "3.5.9": {
        "name": "Temporary Password",
        "text": "Allow temporary password use for system logons with an immediate change to a permanent password.",
        "family": "IA"
    },
    "3.5.10": {
        "name": "Cryptographic Authentication",
        "text": "Store and transmit only cryptographically-protected passwords.",
        "family": "IA"
    },
    "3.5.11": {
        "name": "Obscure Feedback",
        "text": "Obscure feedback of authentication information.",
        "family": "IA"
    },

    # 3.6 INCIDENT RESPONSE (IR)
    "3.6.1": {
        "name": "Incident Response Plan",
        "text": "Establish an operational incident-handling capability for organizational systems that includes preparation, detection, analysis, containment, recovery, and user response activities.",
        "family": "IR"
    },
    "3.6.2": {
        "name": "Incident Tracking",
        "text": "Track, document, and report incidents to designated officials and/or authorities both internal and external to the organization.",
        "family": "IR"
    },
    "3.6.3": {
        "name": "Incident Testing",
        "text": "Test the organizational incident response capability.",
        "family": "IR"
    },

    # 3.7 MAINTENANCE (MA)
    "3.7.1": {
        "name": "Maintenance Policy",
        "text": "Perform maintenance on organizational systems.",
        "family": "MA"
    },
    "3.7.2": {
        "name": "Controlled Maintenance",
        "text": "Provide controls on the tools, techniques, mechanisms, and personnel used to conduct system maintenance.",
        "family": "MA"
    },
    "3.7.3": {
        "name": "Maintenance Tools",
        "text": "Ensure equipment removed for off-site maintenance is sanitized of any CUI.",
        "family": "MA"
    },
    "3.7.4": {
        "name": "Nonlocal Maintenance",
        "text": "Check media containing diagnostic and test programs for malicious code before the media are used in organizational systems.",
        "family": "MA"
    },
    "3.7.5": {
        "name": "Maintenance Personnel",
        "text": "Require multifactor authentication to establish nonlocal maintenance sessions via external network connections and terminate such connections when nonlocal maintenance is complete.",
        "family": "MA"
    },
    "3.7.6": {
        "name": "Maintenance Supervision",
        "text": "Supervise the maintenance activities of maintenance personnel without required access authorization.",
        "family": "MA"
    },

    # 3.8 MEDIA PROTECTION (MP)
    "3.8.1": {
        "name": "Media Access",
        "text": "Protect (i.e., physically control and securely store) system media containing CUI, both paper and digital.",
        "family": "MP"
    },
    "3.8.2": {
        "name": "Media Disposal",
        "text": "Limit access to CUI on system media to authorized users.",
        "family": "MP"
    },
    "3.8.3": {
        "name": "Media Sanitization",
        "text": "Sanitize or destroy system media containing CUI before disposal or release for reuse.",
        "family": "MP"
    },
    "3.8.4": {
        "name": "Media Marking",
        "text": "Mark media with necessary CUI markings and distribution limitations.",
        "family": "MP"
    },
    "3.8.5": {
        "name": "Media Transport",
        "text": "Control access to media containing CUI and maintain accountability for media during transport outside of controlled areas.",
        "family": "MP"
    },
    "3.8.6": {
        "name": "Cryptographic Protection",
        "text": "Implement cryptographic mechanisms to protect the confidentiality of CUI stored on digital media during transport unless otherwise protected by alternative physical safeguards.",
        "family": "MP"
    },
    "3.8.7": {
        "name": "Portable Storage",
        "text": "Control the use of removable media on system components.",
        "family": "MP"
    },
    "3.8.8": {
        "name": "Media Downgrading",
        "text": "Prohibit the use of portable storage devices when such devices have no identifiable owner.",
        "family": "MP"
    },
    "3.8.9": {
        "name": "Media Protection",
        "text": "Protect the confidentiality of backup CUI at storage locations.",
        "family": "MP"
    },

    # 3.9 PERSONNEL SECURITY (PS)
    "3.9.1": {
        "name": "Personnel Screening",
        "text": "Screen individuals prior to authorizing access to organizational systems containing CUI.",
        "family": "PS"
    },
    "3.9.2": {
        "name": "Personnel Termination",
        "text": "Ensure that organizational systems containing CUI are protected during and after personnel actions such as terminations and transfers.",
        "family": "PS"
    },

    # 3.10 PHYSICAL PROTECTION (PE)
    "3.10.1": {
        "name": "Physical Access Authorizations",
        "text": "Limit physical access to organizational systems, equipment, and the respective operating environments to authorized individuals.",
        "family": "PE"
    },
    "3.10.2": {
        "name": "Physical Access Controls",
        "text": "Protect and monitor the physical facility and support infrastructure for organizational systems.",
        "family": "PE"
    },
    "3.10.3": {
        "name": "Escort Visitors",
        "text": "Escort visitors and monitor visitor activity.",
        "family": "PE"
    },
    "3.10.4": {
        "name": "Physical Access Logs",
        "text": "Maintain audit logs of physical access.",
        "family": "PE"
    },
    "3.10.5": {
        "name": "Physical Access Devices",
        "text": "Control and manage physical access devices.",
        "family": "PE"
    },
    "3.10.6": {
        "name": "Monitoring Physical Access",
        "text": "Enforce safeguarding measures for CUI at alternate work sites.",
        "family": "PE"
    },

    # 3.11 RISK ASSESSMENT (RA)
    "3.11.1": {
        "name": "Risk Assessment",
        "text": "Periodically assess the risk to organizational operations (including mission, functions, image, or reputation), organizational assets, and individuals, resulting from the operation of organizational systems and the associated processing, storage, or transmission of CUI.",
        "family": "RA"
    },
    "3.11.2": {
        "name": "Vulnerability Scanning",
        "text": "Scan for vulnerabilities in organizational systems and applications periodically and when new vulnerabilities affecting those systems and applications are identified.",
        "family": "RA"
    },
    "3.11.3": {
        "name": "Remediation",
        "text": "Remediate vulnerabilities in accordance with risk assessments.",
        "family": "RA"
    },

    # 3.12 SECURITY ASSESSMENT (CA)
    "3.12.1": {
        "name": "Security Assessments",
        "text": "Periodically assess the security controls in organizational systems to determine if the controls are effective in their application.",
        "family": "CA"
    },
    "3.12.2": {
        "name": "Plans of Action",
        "text": "Develop and implement plans of action designed to correct deficiencies and reduce or eliminate vulnerabilities in organizational systems.",
        "family": "CA"
    },
    "3.12.3": {
        "name": "System Interconnections",
        "text": "Monitor security controls on an ongoing basis to ensure the continued effectiveness of the controls.",
        "family": "CA"
    },
    "3.12.4": {
        "name": "Security Control Testing",
        "text": "Develop, document, and periodically update system security plans that describe system boundaries, system environments of operation, how security requirements are implemented, and the relationships with or connections to other systems.",
        "family": "CA"
    },

    # 3.13 SYSTEM AND COMMUNICATIONS PROTECTION (SC)
    "3.13.1": {
        "name": "Boundary Protection",
        "text": "Monitor, control, and protect communications (i.e., information transmitted or received by organizational systems) at the external boundaries and key internal boundaries of organizational systems.",
        "family": "SC"
    },
    "3.13.2": {
        "name": "Application Partitioning",
        "text": "Employ architectural designs, software development techniques, and systems engineering principles that promote effective information security within organizational systems.",
        "family": "SC"
    },
    "3.13.3": {
        "name": "Security Function Isolation",
        "text": "Separate user functionality from system management functionality.",
        "family": "SC"
    },
    "3.13.4": {
        "name": "Information in Shared Resources",
        "text": "Prevent unauthorized and unintended information transfer via shared system resources.",
        "family": "SC"
    },
    "3.13.5": {
        "name": "Denial of Service Protection",
        "text": "Implement subnetworks for publicly accessible system components that are physically or logically separated from internal networks.",
        "family": "SC"
    },
    "3.13.6": {
        "name": "Network Segmentation",
        "text": "Deny network communications traffic by default and allow network communications traffic by exception (i.e., deny all, permit by exception).",
        "family": "SC"
    },
    "3.13.7": {
        "name": "Split Tunneling",
        "text": "Prevent remote devices from simultaneously establishing non-remote connections with organizational systems and communicating via some other connection to resources in external networks (i.e., split tunneling).",
        "family": "SC"
    },
    "3.13.8": {
        "name": "Cryptographic Protection",
        "text": "Implement cryptographic mechanisms to prevent unauthorized disclosure of CUI during transmission unless otherwise protected by alternative physical safeguards.",
        "family": "SC"
    },
    "3.13.9": {
        "name": "Session Termination",
        "text": "Terminate network connections associated with communications sessions at the end of the sessions or after a defined period of inactivity.",
        "family": "SC"
    },
    "3.13.10": {
        "name": "Cryptographic Key Management",
        "text": "Establish and manage cryptographic keys for cryptography employed in organizational systems.",
        "family": "SC"
    },
    "3.13.11": {
        "name": "CUI Confidentiality",
        "text": "Employ FIPS-validated cryptography when used to protect the confidentiality of CUI.",
        "family": "SC"
    },
    "3.13.12": {
        "name": "Collaborative Computing Devices",
        "text": "Prohibit remote activation of collaborative computing devices and provide indication of devices in use to users present at the device.",
        "family": "SC"
    },
    "3.13.13": {
        "name": "Mobile Code",
        "text": "Control and monitor the use of mobile code.",
        "family": "SC"
    },
    "3.13.14": {
        "name": "Voice over Internet Protocol",
        "text": "Control and monitor the use of Voice over Internet Protocol (VoIP) technologies.",
        "family": "SC"
    },
    "3.13.15": {
        "name": "Authenticity Protection",
        "text": "Protect the authenticity of communications sessions.",
        "family": "SC"
    },
    "3.13.16": {
        "name": "Transmission Confidentiality",
        "text": "Protect the confidentiality of CUI at rest.",
        "family": "SC"
    },

    # 3.14 SYSTEM AND INFORMATION INTEGRITY (SI)
    "3.14.1": {
        "name": "Flaw Remediation",
        "text": "Identify, report, and correct system flaws in a timely manner.",
        "family": "SI"
    },
    "3.14.2": {
        "name": "Malicious Code Protection",
        "text": "Provide protection from malicious code at designated locations within organizational systems.",
        "family": "SI"
    },
    "3.14.3": {
        "name": "Security Alerts and Advisories",
        "text": "Monitor system security alerts and advisories and take action in response.",
        "family": "SI"
    },
    "3.14.4": {
        "name": "Software and Firmware Integrity",
        "text": "Update malicious code protection mechanisms when new releases are available.",
        "family": "SI"
    },
    "3.14.5": {
        "name": "Spam Protection",
        "text": "Perform periodic scans of organizational systems and real-time scans of files from external sources as files are downloaded, opened, or executed.",
        "family": "SI"
    },
    "3.14.6": {
        "name": "Information Handling and Retention",
        "text": "Monitor organizational systems, including inbound and outbound communications traffic, to detect attacks and indicators of potential attacks.",
        "family": "SI"
    },
    "3.14.7": {
        "name": "Information System Monitoring",
        "text": "Identify unauthorized use of organizational systems.",
        "family": "SI"
        }
}


class NIST80171ControlsRev3:
    """NIST SP 800-171 Rev 3 Control Definitions (May 2024)"""

    CONTROLS = {
    # 03.01 ACCESS CONTROL (AC) - 22 controls
    "03.01.01": {
        "name": "Account Management",
        "text": "Limit system access to authorized users, processes acting on behalf of authorized users, and devices (including other systems).",
        "family": "AC"
    },
    "03.01.02": {
        "name": "Access Enforcement",
        "text": "Limit system access to the types of transactions and functions that authorized users are permitted to execute.",
        "family": "AC"
    },
    "03.01.03": {
        "name": "Information Flow Enforcement",
        "text": "Control the flow of CUI in accordance with approved authorizations.",
        "family": "AC"
    },
    "03.01.04": {
        "name": "Separation of Duties",
        "text": "Separate the duties of individuals to reduce the risk of malevolent activity without collusion.",
        "family": "AC"
    },
    "03.01.05": {
        "name": "Least Privilege",
        "text": "Employ the principle of least privilege, including for specific security functions and privileged accounts.",
        "family": "AC"
    },
    "03.01.06": {
        "name": "Non-Privileged Account Use",
        "text": "Use non-privileged accounts or roles when accessing non-security functions.",
        "family": "AC"
    },
    "03.01.07": {
        "name": "Privileged Function Execution",
        "text": "Prevent non-privileged users from executing privileged functions and capture the execution of such functions in audit logs.",
        "family": "AC"
    },
    "03.01.08": {
        "name": "Unsuccessful Logon Attempts",
        "text": "Limit unsuccessful logon attempts.",
        "family": "AC"
    },
    "03.01.09": {
        "name": "System Use Notification",
        "text": "Provide privacy and security notices consistent with CUI rules.",
        "family": "AC"
    },
    "03.01.10": {
        "name": "Session Lock",
        "text": "Use session lock with pattern-hiding displays after a period of inactivity.",
        "family": "AC"
    },
    "03.01.11": {
        "name": "Session Termination",
        "text": "Terminate (automatically) a user session after a defined condition.",
        "family": "AC"
    },
    "03.01.12": {
        "name": "Remote Access Management",
        "text": "Monitor and control remote access sessions.",
        "family": "AC"
    },
    "03.01.13": {
        "name": "Remote Access Confidentiality",
        "text": "Employ cryptographic mechanisms to protect the confidentiality of remote access sessions.",
        "family": "AC"
    },
    "03.01.14": {
        "name": "Remote Access Routing",
        "text": "Route remote access via managed access control points.",
        "family": "AC"
    },
    "03.01.15": {
        "name": "Privileged Remote Access",
        "text": "Authorize remote execution of privileged commands and access to security-relevant information via remote access only for documented operational needs.",
        "family": "AC"
    },
    "03.01.16": {
        "name": "Wireless Access Authorization",
        "text": "Authorize wireless access prior to allowing such connections.",
        "family": "AC"
    },
    "03.01.17": {
        "name": "Wireless Access Protection",
        "text": "Protect wireless access using authentication and encryption.",
        "family": "AC"
    },
    "03.01.18": {
        "name": "Mobile Device Management",
        "text": "Control connection of mobile devices.",
        "family": "AC"
    },
    "03.01.19": {
        "name": "Encryption of CUI on Mobile",
        "text": "Encrypt CUI on mobile devices and mobile computing platforms.",
        "family": "AC"
    },
    "03.01.20": {
        "name": "External System Connections",
        "text": "Verify and control/limit connections to external systems.",
        "family": "AC"
    },
    "03.01.21": {
        "name": "Portable Storage Device Use",
        "text": "Limit use of portable storage devices on external systems.",
        "family": "AC"
    },
    "03.01.22": {
        "name": "Publicly Accessible Content",
        "text": "Control CUI posted or processed on publicly accessible systems.",
        "family": "AC"
    },

    # 03.02 AWARENESS AND TRAINING (AT) - 3 controls
    "03.02.01": {
        "name": "Literacy Training and Awareness",
        "text": "Provide security literacy training to organizational personnel.",
        "family": "AT"
    },
    "03.02.02": {
        "name": "Role-Based Training",
        "text": "Provide role-based security training to personnel with assigned security roles and responsibilities.",
        "family": "AT"
    },
    "03.02.03": {
        "name": "Insider Threat Awareness",
        "text": "Provide security awareness training on recognizing and reporting potential indicators of insider threat.",
        "family": "AT"
    },

    # 03.03 AUDIT AND ACCOUNTABILITY (AU) - 9 controls
    "03.03.01": {
        "name": "Event Logging",
        "text": "Create and retain system audit logs and records to the extent needed to enable the monitoring, analysis, investigation, and reporting of unlawful or unauthorized system activity.",
        "family": "AU"
    },
    "03.03.02": {
        "name": "User Accountability",
        "text": "Ensure that the actions of individual system users can be uniquely traced to those users so they can be held accountable for their actions.",
        "family": "AU"
    },
    "03.03.03": {
        "name": "Event Review",
        "text": "Review and update logged events.",
        "family": "AU"
    },
    "03.03.04": {
        "name": "Audit Failure Alerting",
        "text": "Alert in the event of an audit logging process failure.",
        "family": "AU"
    },
    "03.03.05": {
        "name": "Audit Correlation",
        "text": "Correlate audit record review, analysis, and reporting processes for investigation and response to indications of unlawful, unauthorized, suspicious, or unusual activity.",
        "family": "AU"
    },
    "03.03.06": {
        "name": "Reduction and Report Generation",
        "text": "Provide audit record reduction and report generation to support on-demand analysis and reporting.",
        "family": "AU"
    },
    "03.03.07": {
        "name": "Authoritative Time Source",
        "text": "Provide a system capability that compares and synchronizes internal system clocks with an authoritative source to generate time stamps for audit records.",
        "family": "AU"
    },
    "03.03.08": {
        "name": "Audit Record Protection",
        "text": "Protect audit information and audit logging tools from unauthorized access, modification, and deletion.",
        "family": "AU"
    },
    "03.03.09": {
        "name": "Audit Management",
        "text": "Limit management of audit logging functionality to a subset of privileged users.",
        "family": "AU"
    },

    # 03.04 CONFIGURATION MANAGEMENT (CM) - 12 controls
    "03.04.01": {
        "name": "Baseline Configuration",
        "text": "Establish and maintain baseline configurations and inventories of organizational systems throughout the respective system development life cycles.",
        "family": "CM"
    },
    "03.04.02": {
        "name": "Security Configuration Settings",
        "text": "Establish and enforce security configuration settings for information technology products employed in organizational systems.",
        "family": "CM"
    },
    "03.04.03": {
        "name": "System Change Control",
        "text": "Track, review, approve or disapprove, and log changes to organizational systems.",
        "family": "CM"
    },
    "03.04.04": {
        "name": "Security Impact Analysis",
        "text": "Analyze the security impact of changes prior to implementation.",
        "family": "CM"
    },
    "03.04.05": {
        "name": "Access Restrictions for Change",
        "text": "Define, document, approve, and enforce physical and logical access restrictions associated with changes to organizational systems.",
        "family": "CM"
    },
    "03.04.06": {
        "name": "Least Functionality",
        "text": "Employ the principle of least functionality by configuring organizational systems to provide only essential capabilities.",
        "family": "CM"
    },
    "03.04.07": {
        "name": "Nonessential Functionality",
        "text": "Restrict, disable, or prevent the use of nonessential programs, functions, ports, protocols, and services.",
        "family": "CM"
    },
    "03.04.08": {
        "name": "Application Execution Policy",
        "text": "Apply deny-by-exception or deny-all, permit-by-exception policy to prevent the use of unauthorized software.",
        "family": "CM"
    },
    "03.04.09": {
        "name": "User-Installed Software",
        "text": "Control and monitor user-installed software.",
        "family": "CM"
    },
    "03.04.10": {
        "name": "System Component Inventory",
        "text": "Develop and document an inventory of system components.",
        "family": "CM"
    },
    "03.04.11": {
        "name": "Information Location",
        "text": "Identify where CUI is processed, stored, and transmitted.",
        "family": "CM"
    },
    "03.04.12": {
        "name": "Automated Configuration Management",
        "text": "Employ automated mechanisms to maintain an up-to-date, complete, accurate, and readily available baseline configuration of the system.",
        "family": "CM"
    },

    # 03.05 IDENTIFICATION AND AUTHENTICATION (IA) - 12 controls
    "03.05.01": {
        "name": "User Identification",
        "text": "Identify system users, processes acting on behalf of users, and devices.",
        "family": "IA"
    },
    "03.05.02": {
        "name": "Device Identification and Authentication",
        "text": "Authenticate the identities of users, processes, or devices before allowing access to organizational systems.",
        "family": "IA"
    },
    "03.05.03": {
        "name": "Multifactor Authentication",
        "text": "Use multifactor authentication for local and network access to privileged accounts and for network access to non-privileged accounts.",
        "family": "IA"
    },
    "03.05.04": {
        "name": "Replay-Resistant Authentication",
        "text": "Employ replay-resistant authentication mechanisms for network access to privileged and non-privileged accounts.",
        "family": "IA"
    },
    "03.05.05": {
        "name": "Identifier Reuse",
        "text": "Prevent reuse of identifiers for a defined period.",
        "family": "IA"
    },
    "03.05.06": {
        "name": "Identifier Handling",
        "text": "Disable identifiers after a defined period of inactivity.",
        "family": "IA"
    },
    "03.05.07": {
        "name": "Password Management",
        "text": "Enforce a minimum password complexity and change of characters when new passwords are created.",
        "family": "IA"
    },
    "03.05.08": {
        "name": "Password Reuse",
        "text": "Prohibit password reuse for a specified number of generations.",
        "family": "IA"
    },
    "03.05.09": {
        "name": "Temporary Passwords",
        "text": "Allow temporary password use for system logons with an immediate change to a permanent password.",
        "family": "IA"
    },
    "03.05.10": {
        "name": "Cryptographically-Protected Passwords",
        "text": "Store and transmit only cryptographically-protected passwords.",
        "family": "IA"
    },
    "03.05.11": {
        "name": "Obscure Feedback",
        "text": "Obscure feedback of authentication information.",
        "family": "IA"
    },
    "03.05.12": {
        "name": "Authenticator Management",
        "text": "Manage system authenticators by verifying identity of the user, device, or service before distributing, establishing, or changing authenticators.",
        "family": "IA"
    },

    # 03.06 INCIDENT RESPONSE (IR) - 4 controls
    "03.06.01": {
        "name": "Incident Handling",
        "text": "Establish an operational incident-handling capability for organizational systems that includes preparation, detection, analysis, containment, recovery, and user response activities.",
        "family": "IR"
    },
    "03.06.02": {
        "name": "Incident Reporting",
        "text": "Track, document, and report incidents to designated officials and/or authorities both internal and external to the organization.",
        "family": "IR"
    },
    "03.06.03": {
        "name": "Incident Response Testing",
        "text": "Test the organizational incident response capability.",
        "family": "IR"
    },
    "03.06.04": {
        "name": "Incident Response Training",
        "text": "Provide incident response training to system users consistent with assigned roles and responsibilities.",
        "family": "IR"
    },

    # 03.07 MAINTENANCE (MA) - 6 controls
    "03.07.01": {
        "name": "Controlled Maintenance",
        "text": "Perform maintenance on organizational systems.",
        "family": "MA"
    },
    "03.07.02": {
        "name": "System Maintenance Controls",
        "text": "Provide controls on the tools, techniques, mechanisms, and personnel used to conduct system maintenance.",
        "family": "MA"
    },
    "03.07.03": {
        "name": "Equipment Sanitization",
        "text": "Ensure equipment removed for off-site maintenance is sanitized of any CUI.",
        "family": "MA"
    },
    "03.07.04": {
        "name": "Media Inspection",
        "text": "Check media containing diagnostic and test programs for malicious code before the media are used in organizational systems.",
        "family": "MA"
    },
    "03.07.05": {
        "name": "Nonlocal Maintenance",
        "text": "Require multifactor authentication to establish nonlocal maintenance sessions and terminate such connections when nonlocal maintenance is complete.",
        "family": "MA"
    },
    "03.07.06": {
        "name": "Maintenance Personnel",
        "text": "Supervise the maintenance activities of maintenance personnel without required access authorization.",
        "family": "MA"
    },

    # 03.08 MEDIA PROTECTION (MP) - 9 controls
    "03.08.01": {
        "name": "Media Access",
        "text": "Protect system media containing CUI, both paper and digital.",
        "family": "MP"
    },
    "03.08.02": {
        "name": "Media Marking",
        "text": "Mark media with necessary CUI markings and distribution limitations.",
        "family": "MP"
    },
    "03.08.03": {
        "name": "Media Storage",
        "text": "Control access to media containing CUI and maintain accountability for media during transport outside of controlled areas.",
        "family": "MP"
    },
    "03.08.04": {
        "name": "Media Transport",
        "text": "Control the use of removable media on system components.",
        "family": "MP"
    },
    "03.08.05": {
        "name": "Media Sanitization",
        "text": "Sanitize or destroy system media containing CUI before disposal or release for reuse.",
        "family": "MP"
    },
    "03.08.06": {
        "name": "Media Use",
        "text": "Implement cryptographic mechanisms to protect the confidentiality of CUI stored on digital media during transport.",
        "family": "MP"
    },
    "03.08.07": {
        "name": "Removable Media",
        "text": "Control the use of removable media on system components.",
        "family": "MP"
    },
    "03.08.08": {
        "name": "Shared Media",
        "text": "Prohibit the use of portable storage devices when such devices have no identifiable owner.",
        "family": "MP"
    },
    "03.08.09": {
        "name": "Protect Backups",
        "text": "Protect the confidentiality of backup CUI at storage locations.",
        "family": "MP"
    },

    # 03.09 PERSONNEL SECURITY (PS) - 2 controls
    "03.09.01": {
        "name": "Personnel Screening",
        "text": "Screen individuals prior to authorizing access to organizational systems containing CUI.",
        "family": "PS"
    },
    "03.09.02": {
        "name": "Personnel Termination and Transfer",
        "text": "Ensure that CUI and organizational systems are protected during and after personnel actions such as terminations and transfers.",
        "family": "PS"
    },

    # 03.10 PHYSICAL PROTECTION (PE) - 6 controls
    "03.10.01": {
        "name": "Physical Access Authorization",
        "text": "Limit physical access to organizational systems, equipment, and the respective operating environments to authorized individuals.",
        "family": "PE"
    },
    "03.10.02": {
        "name": "Physical Access Control",
        "text": "Protect and monitor the physical facility and support infrastructure for organizational systems.",
        "family": "PE"
    },
    "03.10.03": {
        "name": "Escort Visitors",
        "text": "Escort visitors and monitor visitor activity.",
        "family": "PE"
    },
    "03.10.04": {
        "name": "Physical Access Log",
        "text": "Maintain audit logs of physical access.",
        "family": "PE"
    },
    "03.10.05": {
        "name": "Manage Physical Access",
        "text": "Control and manage physical access devices.",
        "family": "PE"
    },
    "03.10.06": {
        "name": "Alternative Work Sites",
        "text": "Enforce safeguards to protect CUI at alternate work sites.",
        "family": "PE"
    },

    # 03.11 RISK ASSESSMENT (RA) - 10 controls (expanded from 3 in Rev 2)
    "03.11.01": {
        "name": "Risk Assessment",
        "text": "Periodically assess the risk to organizational operations, assets, and individuals resulting from the operation of organizational systems and the associated processing, storage, or transmission of CUI.",
        "family": "RA"
    },
    "03.11.02": {
        "name": "Vulnerability Monitoring and Scanning",
        "text": "Scan for vulnerabilities in organizational systems periodically and when new vulnerabilities affecting those systems are identified.",
        "family": "RA"
    },
    "03.11.03": {
        "name": "Vulnerability Remediation",
        "text": "Remediate vulnerabilities in accordance with risk assessments.",
        "family": "RA"
    },
    "03.11.04": {
        "name": "Risk Response",
        "text": "Respond to findings from security assessments, monitoring, and audits.",
        "family": "RA"
    },
    "03.11.05": {
        "name": "Threat Intelligence",
        "text": "Use threat intelligence to inform risk assessments, response actions, and security requirements.",
        "family": "RA"
    },
    "03.11.06": {
        "name": "Cybersecurity Supply Chain Risk",
        "text": "Identify, assess, and manage cybersecurity supply chain risks associated with the development, acquisition, maintenance, and disposal of systems.",
        "family": "RA"
    },
    "03.11.07": {
        "name": "Operational Resilience",
        "text": "Develop and implement plans for recovering from adverse events affecting organizational systems.",
        "family": "RA"
    },
    "03.11.08": {
        "name": "Business Impact Analysis",
        "text": "Conduct a business impact analysis to characterize the consequences of incidents on organizational missions and business functions.",
        "family": "RA"
    },
    "03.11.09": {
        "name": "Criticality Analysis",
        "text": "Identify critical system components and functions by performing a criticality analysis.",
        "family": "RA"
    },
    "03.11.10": {
        "name": "Data Classification",
        "text": "Identify and classify CUI and establish handling requirements consistent with applicable law, Executive Orders, directives, and regulations.",
        "family": "RA"
    },

    # 03.12 SECURITY ASSESSMENT (CA) - 6 controls (expanded from 4 in Rev 2)
    "03.12.01": {
        "name": "Security Control Assessment",
        "text": "Periodically assess the security controls in organizational systems to determine if the controls are effective in their application.",
        "family": "CA"
    },
    "03.12.02": {
        "name": "Plan of Action",
        "text": "Develop and implement plans of action designed to correct deficiencies and reduce or eliminate vulnerabilities in organizational systems.",
        "family": "CA"
    },
    "03.12.03": {
        "name": "Security Monitoring",
        "text": "Monitor security controls on an ongoing basis to ensure the continued effectiveness of the controls.",
        "family": "CA"
    },
    "03.12.04": {
        "name": "System Security Plan",
        "text": "Develop, document, and periodically update system security plans that describe system boundaries, system environments of operation, how security requirements are implemented.",
        "family": "CA"
    },
    "03.12.05": {
        "name": "Security Authorization",
        "text": "Authorize the operation of organizational systems and any associated system connections.",
        "family": "CA"
    },
    "03.12.06": {
        "name": "External System Connections",
        "text": "Verify that the security requirements are satisfied before authorizing connections to external systems.",
        "family": "CA"
    },

    # 03.13 SYSTEM AND COMMUNICATIONS PROTECTION (SC) - 16 controls
    "03.13.01": {
        "name": "Boundary Protection",
        "text": "Monitor, control, and protect communications at the external boundaries and key internal boundaries of organizational systems.",
        "family": "SC"
    },
    "03.13.02": {
        "name": "Security Engineering Principles",
        "text": "Employ security engineering principles in the specification, design, development, implementation, and modification of organizational systems.",
        "family": "SC"
    },
    "03.13.03": {
        "name": "Role Separation",
        "text": "Separate user functionality from system management functionality.",
        "family": "SC"
    },
    "03.13.04": {
        "name": "Shared Resource Control",
        "text": "Prevent unauthorized and unintended information transfer via shared system resources.",
        "family": "SC"
    },
    "03.13.05": {
        "name": "Public Access Protection",
        "text": "Implement subnetworks for publicly accessible system components that are physically or logically separated from internal networks.",
        "family": "SC"
    },
    "03.13.06": {
        "name": "Network Communication Denial",
        "text": "Deny network communications traffic by default and allow network communications traffic by exception.",
        "family": "SC"
    },
    "03.13.07": {
        "name": "Split Tunneling",
        "text": "Prevent remote devices from simultaneously establishing non-remote connections with the system and communicating via some other connection to resources in external networks.",
        "family": "SC"
    },
    "03.13.08": {
        "name": "Data in Transit",
        "text": "Implement cryptographic mechanisms to prevent unauthorized disclosure of CUI during transmission unless otherwise protected by alternative physical safeguards.",
        "family": "SC"
    },
    "03.13.09": {
        "name": "Network Disconnect",
        "text": "Terminate network connections associated with communications sessions at the end of the sessions or after a defined period of inactivity.",
        "family": "SC"
    },
    "03.13.10": {
        "name": "Key Management",
        "text": "Establish and manage cryptographic keys for cryptography employed in organizational systems.",
        "family": "SC"
    },
    "03.13.11": {
        "name": "CUI Encryption",
        "text": "Employ FIPS-validated cryptography when used to protect the confidentiality of CUI.",
        "family": "SC"
    },
    "03.13.12": {
        "name": "Collaborative Device Control",
        "text": "Prohibit remote activation of collaborative computing devices and provide indication of use to users present at the device.",
        "family": "SC"
    },
    "03.13.13": {
        "name": "Mobile Code",
        "text": "Control and monitor the use of mobile code.",
        "family": "SC"
    },
    "03.13.14": {
        "name": "VoIP Technologies",
        "text": "Control and monitor the use of Voice over Internet Protocol (VoIP) technologies.",
        "family": "SC"
    },
    "03.13.15": {
        "name": "Communications Authenticity",
        "text": "Protect the authenticity of communications sessions.",
        "family": "SC"
    },
    "03.13.16": {
        "name": "Data at Rest",
        "text": "Protect the confidentiality of CUI at rest.",
        "family": "SC"
    },

    # 03.14 SYSTEM AND INFORMATION INTEGRITY (SI) - 9 controls (expanded from 7 in Rev 2)
    "03.14.01": {
        "name": "Flaw Remediation",
        "text": "Identify, report, and correct system flaws in a timely manner.",
        "family": "SI"
    },
    "03.14.02": {
        "name": "Malicious Code Protection",
        "text": "Provide protection from malicious code at designated locations within organizational systems.",
        "family": "SI"
    },
    "03.14.03": {
        "name": "Security Alert Monitoring",
        "text": "Monitor system security alerts and advisories and take action in response.",
        "family": "SI"
    },
    "03.14.04": {
        "name": "Update Malicious Code Protection",
        "text": "Update malicious code protection mechanisms when new releases are available.",
        "family": "SI"
    },
    "03.14.05": {
        "name": "System Monitoring",
        "text": "Monitor systems to detect attacks and indicators of potential attacks and identify unauthorized use of organizational systems.",
        "family": "SI"
    },
    "03.14.06": {
        "name": "Security Monitoring of Systems",
        "text": "Monitor organizational systems, including inbound and outbound communications traffic, to detect attacks and indicators of potential attacks.",
        "family": "SI"
    },
    "03.14.07": {
        "name": "Software and Firmware Integrity",
        "text": "Employ integrity verification tools to detect unauthorized changes to software, firmware, and information.",
        "family": "SI"
    },
    "03.14.08": {
        "name": "Spam Protection",
        "text": "Implement spam protection mechanisms at system entry and exit points.",
        "family": "SI"
    },
    "03.14.09": {
        "name": "Information Input Validation",
        "text": "Check the validity of information inputs to organizational systems.",
        "family": "SI"
    },

    # 03.15 PLANNING (PL) - 2 controls (NEW in Rev 3)
    "03.15.01": {
        "name": "System Security Plan",
        "text": "Develop, document, and disseminate a system security plan that describes the system boundary, the system environment of operation, how security requirements are implemented.",
        "family": "PL"
    },
    "03.15.02": {
        "name": "Rules of Behavior",
        "text": "Establish and provide rules that describe the responsibilities and expected behavior of individuals accessing organizational systems.",
        "family": "PL"
    },

    # 03.16 SUPPLY CHAIN RISK MANAGEMENT (SR) - 3 controls (NEW in Rev 3)
    "03.16.01": {
        "name": "Supply Chain Risk Management Plan",
        "text": "Develop a plan for managing supply chain risks associated with the research and development, design, manufacturing, acquisition, delivery, integration, operations, maintenance, and disposal of systems.",
        "family": "SR"
    },
    "03.16.02": {
        "name": "Supplier Review",
        "text": "Use due diligence when selecting suppliers and regularly review supplier risk.",
        "family": "SR"
    },
    "03.16.03": {
        "name": "Supply Chain Controls and Processes",
        "text": "Employ cybersecurity controls and processes to address risks associated with the supply chain.",
        "family": "SR"
    },

    # 03.17 PROGRAM MANAGEMENT (PM) - 1 control (NEW in Rev 3)
    "03.17.01": {
        "name": "Information Security Program",
        "text": "Develop and implement an organization-wide program to manage information security risk to organizational operations and assets.",
        "family": "PM"
    },
}


def select_nist_version() -> tuple:
    """Interactive menu to select NIST SP 800-171 revision.
    Returns (version_label, controls_dict, standard_label).
    """
    print("\n" + "=" * 60)
    print("  NIST SP 800-171 VERSION SELECTION")
    print("=" * 60)
    print("  [1]  Rev 2  (110 controls, 14 families) - Original")
    print("  [2]  Rev 3  (132 controls, 17 families) - May 2024")
    print("=" * 60)
    while True:
        choice = input("  Select version [1/2]: ").strip()
        if choice == "1":
            return ("Rev 2", NIST80171Controls.CONTROLS, "NIST SP 800-171 Rev 2")
        if choice == "2":
            return ("Rev 3", NIST80171ControlsRev3.CONTROLS, "NIST SP 800-171 Rev 3")
        print("  Invalid choice. Enter 1 or 2.")


class NetworkScanner:
    """Network and endpoint scanner using nmap"""

    def __init__(self):
        try:
            self.nm = nmap.PortScanner()
            self.topology = None
        except nmap.PortScannerError as e:
            logger.error(f"Nmap initialization failed: {e}")
            logger.error("Please install nmap:")
            logger.error("Windows: Download from https://nmap.org/download.html")
            logger.error("Linux: sudo apt-get install nmap")
            logger.error("macOS: brew install nmap")
            raise SystemExit("Nmap is required but not found. Please install nmap and try again.")

    def scan_network(self, network_range: str) -> List[SystemInfo]:
        """Scan network range for active hosts and services"""
        logger.info(f"[SCAN] Scanning network range: {network_range}")
        _vsection(f"NETWORK DISCOVERY  |  Range: {network_range}")
        _vprint(f"[>] Running host discovery (ping sweep)...")
        systems = []

        try:
            is_windows = platform.system().lower() == 'windows'

            if is_windows:
                scan_args = '-sn'
            else:
                scan_args = '-sn -PR -PS21,22,23,25,53,80,110,111,135,139,143,443,993,995'

            self.nm.scan(hosts=network_range, arguments=scan_args)
            active_hosts = [host for host in self.nm.all_hosts() if self.nm[host].state() == 'up']

            logger.info(f"[SCAN] Found {len(active_hosts)} active hosts")
            if _verbose_mode:
                _vprint(f"[+] Found {len(active_hosts)} active host(s):")
                for h in active_hosts:
                    _vprint(f"      {h}  (up)")

            # Full-depth scan flags used for every discovered host
            DEEP_SCAN_ARGS = '-A -vvv -T3 -p 1-65535 --script "default or discovery or broadcast"'

            for host in active_hosts:
                try:
                    logger.info(f"[SCAN] Deep scanning host {host} ...")
                    _vprint(f"\n  [>] Deep scanning: {host}")
                    self.nm.scan(host, arguments=DEEP_SCAN_ARGS)

                    if host in self.nm.all_hosts():
                        host_info = self.nm[host]

                        os_info = "Unknown"
                        os_version = "Unknown"
                        if host_info.get('osmatch') and len(host_info['osmatch']) > 0:
                            os_match = host_info['osmatch'][0]
                            os_info = os_match.get('name', 'Unknown')
                            if os_match.get('osclass') and len(os_match['osclass']) > 0:
                                os_version = os_match['osclass'][0].get('osfamily', 'Unknown')
                        os_type = os_info

                        open_ports = []
                        services = {}

                        if 'tcp' in host_info:
                            for port, port_info in host_info['tcp'].items():
                                if port_info['state'] == 'open':
                                    open_ports.append(port)
                                    service_name = port_info.get('name', 'unknown')
                                    service_version = port_info.get('version', '')
                                    services[port] = f"{service_name} {service_version}".strip()

                        hostname = host_info.hostname() if host_info.hostname() else host
                        mac_address = None
                        vendor = None

                        if 'addresses' in host_info:
                            addresses = host_info['addresses']
                            if 'mac' in addresses:
                                mac_address = addresses['mac']

                        if 'vendor' in host_info and host_info['vendor']:
                            vendor = list(host_info['vendor'].values())[0]

                        hop_count = None
                        gateway = None
                        if not is_windows and 'traceroute' in host_info:
                            traceroute = host_info['traceroute']
                            if traceroute:
                                hop_count = len(traceroute)
                                gateway = traceroute[0].get('ipaddr')
                        else:
                            gateway = self._get_default_gateway_windows() if is_windows else None

                        system_info = SystemInfo(
                            hostname=hostname,
                            ip_address=host,
                            os_type=os_type,
                            os_version=os_version,
                            open_ports=open_ports,
                            services=services,
                            last_scanned=datetime.datetime.now().isoformat(),
                            mac_address=mac_address,
                            vendor=vendor,
                            hop_count=hop_count,
                            gateway=gateway
                        )

                        systems.append(system_info)
                        logger.info(f"[SCAN] Scanned {host}: {len(open_ports)} open ports, OS: {os_type}")
                        if _verbose_mode:
                            _vprint(f"  [+] {host} scan complete:")
                            _vkv("Hostname",    system_info.hostname, indent=2)
                            _vkv("OS",          system_info.os_type, indent=2)
                            ports_str = ', '.join(map(str, system_info.open_ports)) or "None"
                            _vkv("Open Ports",  ports_str, indent=2)
                            svc_items = list(system_info.services.items())[:6]
                            svc_str = ', '.join(f"{p}/{v}" for p, v in svc_items) or "None"
                            _vkv("Services",    svc_str, indent=2)
                            if system_info.mac_address:
                                _vkv("MAC", f"{system_info.mac_address} ({system_info.vendor or 'Unknown'})", indent=2)

                except Exception as e:
                    logger.error(f"[ERROR] Error scanning host {host}: {e}")
                    continue

        except Exception as e:
            logger.error(f"[ERROR] Error during network scan: {e}")

        return systems

    def _get_default_gateway_windows(self) -> Optional[str]:
        """Get default gateway on Windows systems"""
        try:
            result = subprocess.run(['ipconfig'], capture_output=True, text=True, shell=True)
            lines = result.stdout.split('\n')
            for line in lines:
                if 'Default Gateway' in line and ':' in line:
                    gateway = line.split(':')[1].strip()
                    if gateway and gateway != '':
                        return gateway
        except Exception as e:
            logger.debug(f"Could not determine default gateway: {e}")
        return None

    def discover_topology(self, systems: List[SystemInfo]) -> NetworkTopology:
        """Discover and map network topology from scan results"""
        logger.info("[TOPOLOGY] Analyzing network topology...")

        nodes = {}
        edges = []
        subnets = set()
        gateways = set()

        for system in systems:
            node_info = {
                'hostname': system.hostname,
                'ip': system.ip_address,
                'os_type': system.os_type,
                'open_ports': len(system.open_ports),
                'services': list(system.services.values())[:3],
                'mac_address': system.mac_address,
                'vendor': system.vendor,
                'hop_count': system.hop_count or 1,
                'type': self._classify_node_type(system)
            }
            nodes[system.ip_address] = node_info

            try:
                network = ipaddress.ip_network(f"{system.ip_address}/24", strict=False)
                subnets.add(str(network))
            except Exception:
                pass

            if system.gateway:
                gateways.add(system.gateway)
                if system.gateway != system.ip_address:
                    edges.append((system.ip_address, system.gateway))

        for gateway in gateways:
            if gateway not in nodes:
                nodes[gateway] = {
                    'hostname': f'Gateway-{gateway}',
                    'ip': gateway,
                    'os_type': 'Gateway/Router',
                    'open_ports': 0,
                    'services': ['Routing'],
                    'mac_address': None,
                    'vendor': 'Unknown',
                    'hop_count': 0,
                    'type': 'gateway'
                }

        topology = NetworkTopology(
            nodes=nodes,
            edges=edges,
            subnets=list(subnets),
            gateways=list(gateways)
        )

        self.topology = topology
        return topology

    def _classify_node_type(self, system: SystemInfo) -> str:
        """Classify node type based on services and characteristics"""
        services = [service.lower() for service in system.services.values()]
        open_ports = system.open_ports

        if any('http' in service or 'web' in service for service in services):
            return 'web_server'
        elif 22 in open_ports or any('ssh' in service for service in services):
            return 'server'
        elif 'windows' in system.os_type.lower():
            return 'windows_client'
        elif 'linux' in system.os_type.lower():
            return 'linux_client'

        return 'unknown'

    def create_network_diagram(self, topology: NetworkTopology, output_path: str = "network_topology.png"):
        """Create a visual network topology diagram"""
        logger.info("[DIAGRAM] Generating network topology diagram...")

        try:
            G = nx.Graph()

            for ip, node_info in topology.nodes.items():
                G.add_node(ip, **node_info)

            G.add_edges_from(topology.edges)

            plt.figure(figsize=(16, 12))
            plt.clf()

            node_colors = {
                'gateway': '#FF6B6B',
                'web_server': '#4ECDC4',
                'server': '#45B7D1',
                'windows_client': '#AED6F1',
                'linux_client': '#A9DFBF',
                'unknown': '#D5DBDB'
            }

            pos = nx.spring_layout(G, k=3, iterations=50, seed=42)

            for node_type, color in node_colors.items():
                nodes_of_type = [node for node, data in G.nodes(data=True) if data.get('type') == node_type]
                if nodes_of_type:
                    node_sizes = [1000 + (G.nodes[node].get('open_ports', 0) * 100) for node in nodes_of_type]
                    nx.draw_networkx_nodes(G, pos, nodelist=nodes_of_type,
                                         node_color=color, node_size=node_sizes,
                                         alpha=0.8, edgecolors='black', linewidths=1)

            nx.draw_networkx_edges(G, pos, alpha=0.6)

            labels = {}
            for node, data in G.nodes(data=True):
                hostname = data.get('hostname', node)
                if hostname != node:
                    labels[node] = f"{hostname}\n{node}"
                else:
                    labels[node] = node

            nx.draw_networkx_labels(G, pos, labels, font_size=8, font_weight='bold')

            legend_elements = []
            for node_type, color in node_colors.items():
                if any(data.get('type') == node_type for _, data in G.nodes(data=True)):
                    legend_elements.append(mpatches.Patch(color=color, label=node_type.replace('_', ' ').title()))

            plt.legend(handles=legend_elements, loc='upper left', bbox_to_anchor=(0, 1))

            plt.title("Network Topology Diagram\nNISTify NIST SP 800-171 Compliance Assessment",
                     fontsize=16, fontweight='bold', pad=20)

            info_text = f"Total Nodes: {len(G.nodes())}\nTotal Connections: {len(G.edges())}\n"
            info_text += f"Subnets: {len(topology.subnets)}\nGateways: {len(topology.gateways)}"

            plt.text(0.02, 0.98, info_text, transform=plt.gca().transAxes,
                    verticalalignment='top', bbox=dict(boxstyle='round', facecolor='wheat', alpha=0.8))

            plt.axis('off')
            plt.tight_layout()
            plt.savefig(output_path, dpi=300, bbox_inches='tight',
                       facecolor='white', edgecolor='none')
            plt.close()

            topology.network_diagram_path = output_path
            logger.info(f"[DIAGRAM] Network topology diagram saved: {output_path}")

            return output_path

        except Exception as e:
            logger.error(f"[ERROR] Error creating network diagram: {e}")
            return None

class ComplianceAssessorRev2:
    """NIST SP 800-171 Rev 2 compliance assessor (110 controls, 14 families)"""

    def __init__(self):
        self.controls = NIST80171Controls.CONTROLS
        self._auto = {
            "access_enforcement": "3.1.2",
            "nonessential_ports":  "3.4.7",
            "boundary_protection": "3.13.1",
        }

    def assess_system(self, system: SystemInfo) -> List[ComplianceResult]:
        """Assess a single system against NIST SP 800-171 Rev 2 controls"""
        results = []
        auto_assessed_ids = set(self._auto.values())

        logger.info(f"[ASSESS] [Rev 2] Assessing {system.hostname} ({system.ip_address})")
        _vsubsection(f"COMPLIANCE ASSESSMENT  |  {system.hostname}  ({system.ip_address})  [Rev 2]")
        _vprint(f"{'Control ID':<14}{'Control Name':<38}{'Status':<18}{'Severity'}")
        _vprint("-" * 78)

        # --- Auto-assessed: Access Enforcement (weak/insecure services) ---
        ctrl_ae = self._auto["access_enforcement"]
        weak_services = self._check_weak_services(system)
        if weak_services:
            results.append(ComplianceResult(
                control_id=ctrl_ae,
                control_name=self.controls[ctrl_ae]["name"],
                control_text=self.controls[ctrl_ae]["text"],
                status="non_compliant",
                finding=f"Potentially insecure services detected: {', '.join(weak_services)}",
                remediation="Disable unnecessary services, implement strong authentication, and restrict access",
                severity="high",
                evidence=[f"Open ports: {system.open_ports}", f"Services: {system.services}"]
            ))
            self._vresult(results[-1])
        else:
            results.append(ComplianceResult(
                control_id=ctrl_ae,
                control_name=self.controls[ctrl_ae]["name"],
                control_text=self.controls[ctrl_ae]["text"],
                status="compliant",
                finding="No obviously insecure services detected",
                remediation="Continue monitoring for unauthorized services",
                severity="low",
                evidence=[f"Services reviewed: {list(system.services.values())}"]
            ))
            self._vresult(results[-1])

        # --- Auto-assessed: Nonessential Functionality (unnecessary open ports) ---
        ctrl_np = self._auto["nonessential_ports"]
        unnecessary_ports = self._check_unnecessary_ports(system)
        if unnecessary_ports:
            results.append(ComplianceResult(
                control_id=ctrl_np,
                control_name=self.controls[ctrl_np]["name"],
                control_text=self.controls[ctrl_np]["text"],
                status="non_compliant",
                finding=f"Potentially unnecessary ports open: {unnecessary_ports}",
                remediation="Review and close unnecessary ports, disable unused services",
                severity="medium",
                evidence=[f"Open ports: {system.open_ports}"]
            ))
            self._vresult(results[-1])

        # --- Auto-assessed: Boundary Protection (external-facing services) ---
        ctrl_bp = self._auto["boundary_protection"]
        external_services = self._check_external_services(system)
        if external_services:
            results.append(ComplianceResult(
                control_id=ctrl_bp,
                control_name=self.controls[ctrl_bp]["name"],
                control_text=self.controls[ctrl_bp]["text"],
                status="non_compliant",
                finding=f"External-facing services detected: {external_services}",
                remediation="Implement firewall rules, access controls, and monitoring for external-facing services",
                severity="high",
                evidence=[f"External services: {external_services}"]
            ))
            self._vresult(results[-1])

        # Controls requiring manual review - cannot be automatically assessed via network scan
        manual_review_controls = [
            ("3.1.1", "MANUAL_REVIEW_REQUIRED",
             "Access control policy and user authorization cannot be fully assessed via network scan. Manual review of access control lists and authorization procedures required.",
             "Review and document all authorized users, processes, and devices. Implement role-based access control.",
             "medium"),
            ("3.1.3", "MANUAL_REVIEW_REQUIRED",
             "CUI flow controls cannot be assessed via network scan. Manual review of data flow policies required.",
             "Implement and document CUI flow controls in accordance with approved authorizations.",
             "medium"),
            ("3.1.4", "MANUAL_REVIEW_REQUIRED",
             "Separation of duties cannot be assessed via network scan. Manual review of role assignments required.",
             "Define and enforce separation of duties for sensitive functions and roles.",
             "medium"),
            ("3.1.5", "MANUAL_REVIEW_REQUIRED",
             "Least privilege enforcement cannot be assessed via network scan. Manual review of privilege assignments required.",
             "Review all privileged accounts and apply least privilege principles.",
             "medium"),
            ("3.1.6", "MANUAL_REVIEW_REQUIRED",
             "Non-privileged account usage cannot be assessed via network scan. Manual review required.",
             "Ensure users access non-security functions using non-privileged accounts.",
             "medium"),
            ("3.1.7", "MANUAL_REVIEW_REQUIRED",
             "Privileged function execution controls cannot be assessed via network scan. Manual review required.",
             "Implement controls to prevent non-privileged users from executing privileged functions and enable audit logging.",
             "medium"),
            ("3.1.8", "MANUAL_REVIEW_REQUIRED",
             "Account lockout policies cannot be assessed via network scan. Manual review of authentication configuration required.",
             "Configure account lockout policies to limit unsuccessful logon attempts.",
             "medium"),
            ("3.1.9", "MANUAL_REVIEW_REQUIRED",
             "Privacy and security notices cannot be assessed via network scan. Manual review of login banners required.",
             "Implement privacy and security notices on all system login interfaces.",
             "low"),
            ("3.1.10", "MANUAL_REVIEW_REQUIRED",
             "Session lock policies cannot be assessed via network scan. Manual review of screensaver and session configuration required.",
             "Configure session lock with pattern-hiding displays after a defined period of inactivity.",
             "medium"),
            ("3.1.11", "MANUAL_REVIEW_REQUIRED",
             "Session termination policies cannot be assessed via network scan. Manual review required.",
             "Configure automatic session termination after defined conditions or inactivity periods.",
             "medium"),
            ("3.1.12", "MANUAL_REVIEW_REQUIRED",
             "Remote access monitoring controls cannot be fully assessed via network scan. Manual review required.",
             "Implement monitoring and control mechanisms for all remote access sessions.",
             "medium"),
            ("3.1.13", "MANUAL_REVIEW_REQUIRED",
             "Cryptographic protection of remote access sessions cannot be fully assessed via network scan. Manual review required.",
             "Ensure all remote access sessions use approved cryptographic mechanisms (e.g., TLS, SSH).",
             "high"),
            ("3.1.14", "MANUAL_REVIEW_REQUIRED",
             "Remote access routing controls cannot be assessed via network scan. Manual review required.",
             "Route all remote access through managed access control points.",
             "medium"),
            ("3.1.15", "MANUAL_REVIEW_REQUIRED",
             "Privileged remote access authorization cannot be assessed via network scan. Manual review required.",
             "Document and authorize remote execution of privileged commands and access to security-relevant information.",
             "high"),
            ("3.1.16", "MANUAL_REVIEW_REQUIRED",
             "Wireless access authorization cannot be assessed via network scan. Manual review of wireless configuration required.",
             "Document and authorize all wireless access connections prior to allowing them.",
             "medium"),
            ("3.1.17", "MANUAL_REVIEW_REQUIRED",
             "Wireless access protection cannot be assessed via network scan. Manual review of wireless security configuration required.",
             "Protect wireless access using authentication and encryption (e.g., WPA3, 802.1X).",
             "high"),
            ("3.1.18", "MANUAL_REVIEW_REQUIRED",
             "Mobile device connection controls cannot be assessed via network scan. Manual review of MDM policies required.",
             "Implement mobile device management (MDM) policies and controls.",
             "medium"),
            ("3.1.19", "MANUAL_REVIEW_REQUIRED",
             "CUI encryption on mobile devices cannot be assessed via network scan. Manual review required.",
             "Enforce encryption of CUI on all mobile devices and mobile computing platforms.",
             "high"),
            ("3.1.20", "MANUAL_REVIEW_REQUIRED",
             "External system connection controls cannot be fully assessed via network scan. Manual review required.",
             "Verify, control, and limit connections to and use of external systems.",
             "medium"),
            ("3.1.21", "MANUAL_REVIEW_REQUIRED",
             "Portable storage device controls on external systems cannot be assessed via network scan. Manual review required.",
             "Limit and control use of portable storage devices on external systems.",
             "medium"),
            ("3.1.22", "MANUAL_REVIEW_REQUIRED",
             "Controls for CUI on publicly accessible systems cannot be assessed via network scan. Manual review required.",
             "Review and control CUI posted or processed on publicly accessible systems.",
             "high"),
            ("3.2.1", "MANUAL_REVIEW_REQUIRED",
             "Security awareness program cannot be assessed via network scan. Manual review of training records required.",
             "Establish and maintain a security awareness training program for all system users.",
             "medium"),
            ("3.2.2", "MANUAL_REVIEW_REQUIRED",
             "Security training for assigned duties cannot be assessed via network scan. Manual review of training records required.",
             "Ensure personnel receive role-specific security training for their assigned duties.",
             "medium"),
            ("3.2.3", "MANUAL_REVIEW_REQUIRED",
             "Insider threat awareness training cannot be assessed via network scan. Manual review of training records required.",
             "Provide insider threat awareness training to all personnel.",
             "medium"),
            ("3.3.1", "MANUAL_REVIEW_REQUIRED",
             "Audit log creation and retention policies cannot be fully assessed via network scan. Manual review of logging configuration required.",
             "Configure and maintain audit logging on all systems. Define and enforce log retention policies.",
             "high"),
            ("3.3.2", "MANUAL_REVIEW_REQUIRED",
             "Audit record traceability to individual users cannot be assessed via network scan. Manual review required.",
             "Ensure audit records uniquely identify the user responsible for each action.",
             "high"),
            ("3.3.3", "MANUAL_REVIEW_REQUIRED",
             "Audit record review processes cannot be assessed via network scan. Manual review of procedures required.",
             "Establish and document procedures for periodic review and update of logged events.",
             "medium"),
            ("3.3.4", "MANUAL_REVIEW_REQUIRED",
             "Audit failure alerting cannot be assessed via network scan. Manual review of alerting configuration required.",
             "Configure alerts for audit logging process failures.",
             "high"),
            ("3.3.5", "MANUAL_REVIEW_REQUIRED",
             "Audit correlation processes cannot be assessed via network scan. Manual review of SIEM/log analysis tools required.",
             "Implement audit record correlation capabilities (e.g., SIEM) for investigation and response.",
             "medium"),
            ("3.3.6", "MANUAL_REVIEW_REQUIRED",
             "Audit reduction and report generation cannot be assessed via network scan. Manual review of tools required.",
             "Implement audit reduction and report generation capabilities.",
             "low"),
            ("3.3.7", "MANUAL_REVIEW_REQUIRED",
             "Time synchronization for audit records cannot be fully assessed via network scan. Manual review of NTP configuration required.",
             "Configure all systems to synchronize with an authoritative time source (NTP).",
             "medium"),
            ("3.3.8", "MANUAL_REVIEW_REQUIRED",
             "Audit information protection cannot be assessed via network scan. Manual review of access controls on audit logs required.",
             "Protect audit logs from unauthorized access, modification, and deletion.",
             "high"),
            ("3.3.9", "MANUAL_REVIEW_REQUIRED",
             "Audit management access controls cannot be assessed via network scan. Manual review required.",
             "Limit management of audit logging functionality to privileged users only.",
             "medium"),
            ("3.4.1", "MANUAL_REVIEW_REQUIRED",
             "Baseline configuration documentation cannot be assessed via network scan. Manual review of configuration management records required.",
             "Establish and maintain documented baseline configurations for all systems.",
             "medium"),
            ("3.4.2", "MANUAL_REVIEW_REQUIRED",
             "Security configuration settings enforcement cannot be fully assessed via network scan. Manual review of hardening standards required.",
             "Establish and enforce security configuration settings (e.g., CIS Benchmarks, STIGs).",
             "high"),
            ("3.4.3", "MANUAL_REVIEW_REQUIRED",
             "Configuration change control processes cannot be assessed via network scan. Manual review of change management procedures required.",
             "Implement a configuration change control process for all system changes.",
             "medium"),
            ("3.4.4", "MANUAL_REVIEW_REQUIRED",
             "Security impact analysis cannot be assessed via network scan. Manual review of change management processes required.",
             "Perform security impact analysis for all proposed changes prior to implementation.",
             "medium"),
            ("3.4.5", "MANUAL_REVIEW_REQUIRED",
             "Access restrictions for configuration changes cannot be assessed via network scan. Manual review required.",
             "Define and enforce access restrictions for system configuration changes.",
             "medium"),
            ("3.4.6", "MANUAL_REVIEW_REQUIRED",
             "Least functionality configuration cannot be fully assessed via network scan. Manual review of installed software and enabled services required.",
             "Configure systems to provide only essential capabilities. Remove or disable unnecessary functions.",
             "medium"),
            ("3.4.8", "MANUAL_REVIEW_REQUIRED",
             "Application execution policy cannot be assessed via network scan. Manual review of software whitelisting/blacklisting controls required.",
             "Implement application whitelisting or blacklisting to control software execution.",
             "high"),
            ("3.4.9", "MANUAL_REVIEW_REQUIRED",
             "User-installed software controls cannot be assessed via network scan. Manual review of policies and technical controls required.",
             "Implement controls to monitor and manage user-installed software.",
             "medium"),
            ("3.5.1", "MANUAL_REVIEW_REQUIRED",
             "User identification controls cannot be fully assessed via network scan. Manual review of identity management systems required.",
             "Implement unique identification for all system users, processes, and devices.",
             "high"),
            ("3.5.2", "MANUAL_REVIEW_REQUIRED",
             "Authentication mechanisms cannot be fully assessed via network scan. Manual review of authentication configuration required.",
             "Implement authentication for all users, processes, and devices before granting access.",
             "high"),
            ("3.5.3", "MANUAL_REVIEW_REQUIRED",
             "Multifactor authentication cannot be assessed via network scan. Manual review of authentication systems required.",
             "Implement multifactor authentication for privileged and non-privileged network access.",
             "high"),
            ("3.5.4", "MANUAL_REVIEW_REQUIRED",
             "Replay-resistant authentication mechanisms cannot be assessed via network scan. Manual review required.",
             "Implement replay-resistant authentication (e.g., Kerberos, PKI, OTP).",
             "high"),
            ("3.5.5", "MANUAL_REVIEW_REQUIRED",
             "Identifier reuse prevention cannot be assessed via network scan. Manual review of identity management policies required.",
             "Define and enforce policies preventing reuse of identifiers for a defined period.",
             "medium"),
            ("3.5.6", "MANUAL_REVIEW_REQUIRED",
             "Identifier disablement after inactivity cannot be assessed via network scan. Manual review of account management configuration required.",
             "Configure automatic disabling of identifiers after a defined period of inactivity.",
             "medium"),
            ("3.5.7", "MANUAL_REVIEW_REQUIRED",
             "Password complexity requirements cannot be assessed via network scan. Manual review of password policy configuration required.",
             "Configure and enforce minimum password complexity requirements.",
             "medium"),
            ("3.5.8", "MANUAL_REVIEW_REQUIRED",
             "Password reuse restrictions cannot be assessed via network scan. Manual review of password policy configuration required.",
             "Configure password history to prohibit reuse for a specified number of generations.",
             "medium"),
            ("3.5.9", "MANUAL_REVIEW_REQUIRED",
             "Temporary password controls cannot be assessed via network scan. Manual review of account provisioning processes required.",
             "Implement controls requiring immediate password change after temporary password use.",
             "medium"),
            ("3.5.10", "MANUAL_REVIEW_REQUIRED",
             "Cryptographic password storage and transmission cannot be assessed via network scan. Manual review of authentication systems required.",
             "Ensure passwords are stored and transmitted using approved cryptographic protections.",
             "high"),
            ("3.5.11", "MANUAL_REVIEW_REQUIRED",
             "Authentication feedback obscuring cannot be assessed via network scan. Manual review of authentication interfaces required.",
             "Configure all authentication interfaces to obscure feedback during the authentication process.",
             "medium"),
            ("3.6.1", "MANUAL_REVIEW_REQUIRED",
             "Incident response capability cannot be assessed via network scan. Manual review of incident response plans and procedures required.",
             "Develop and maintain an operational incident response plan covering all phases of incident handling.",
             "high"),
            ("3.6.2", "MANUAL_REVIEW_REQUIRED",
             "Incident tracking and reporting cannot be assessed via network scan. Manual review of incident management processes required.",
             "Implement incident tracking and reporting procedures to designated authorities.",
             "high"),
            ("3.6.3", "MANUAL_REVIEW_REQUIRED",
             "Incident response testing cannot be assessed via network scan. Manual review of exercise records required.",
             "Conduct regular incident response exercises and tabletop drills.",
             "medium"),
            ("3.7.1", "MANUAL_REVIEW_REQUIRED",
             "System maintenance practices cannot be assessed via network scan. Manual review of maintenance schedules and records required.",
             "Establish and maintain documented system maintenance procedures and schedules.",
             "medium"),
            ("3.7.2", "MANUAL_REVIEW_REQUIRED",
             "Controlled maintenance procedures cannot be assessed via network scan. Manual review of maintenance control procedures required.",
             "Implement controls on maintenance tools, techniques, and personnel.",
             "medium"),
            ("3.7.3", "MANUAL_REVIEW_REQUIRED",
             "Maintenance tool sanitization cannot be assessed via network scan. Manual review of sanitization procedures required.",
             "Ensure equipment removed for off-site maintenance is sanitized of CUI before removal.",
             "high"),
            ("3.7.4", "MANUAL_REVIEW_REQUIRED",
             "Media scanning for maintenance programs cannot be assessed via network scan. Manual review of procedures required.",
             "Scan media containing diagnostic and test programs for malicious code before use.",
             "high"),
            ("3.7.5", "MANUAL_REVIEW_REQUIRED",
             "Nonlocal maintenance authentication cannot be assessed via network scan. Manual review of remote maintenance configuration required.",
             "Require MFA for nonlocal maintenance sessions and terminate connections when complete.",
             "high"),
            ("3.7.6", "MANUAL_REVIEW_REQUIRED",
             "Maintenance personnel supervision cannot be assessed via network scan. Manual review of maintenance authorization records required.",
             "Supervise maintenance activities performed by personnel without required authorization.",
             "medium"),
            ("3.8.1", "MANUAL_REVIEW_REQUIRED",
             "Physical media protection cannot be assessed via network scan. Manual review of media handling procedures required.",
             "Implement physical controls to protect and securely store system media containing CUI.",
             "high"),
            ("3.8.2", "MANUAL_REVIEW_REQUIRED",
             "Media access controls cannot be assessed via network scan. Manual review of access control policies for media required.",
             "Limit access to CUI on system media to authorized users.",
             "high"),
            ("3.8.3", "MANUAL_REVIEW_REQUIRED",
             "Media sanitization procedures cannot be assessed via network scan. Manual review of media disposal procedures required.",
             "Implement media sanitization procedures before disposal or reuse.",
             "high"),
            ("3.8.4", "MANUAL_REVIEW_REQUIRED",
             "Media marking practices cannot be assessed via network scan. Manual review of media labeling procedures required.",
             "Mark all media containing CUI with required markings and distribution limitations.",
             "medium"),
            ("3.8.5", "MANUAL_REVIEW_REQUIRED",
             "Media transport controls cannot be assessed via network scan. Manual review of transport procedures required.",
             "Implement controls and accountability measures for media during transport.",
             "high"),
            ("3.8.6", "MANUAL_REVIEW_REQUIRED",
             "Cryptographic protection of media in transport cannot be assessed via network scan. Manual review required.",
             "Implement cryptographic protection for CUI on digital media during transport.",
             "high"),
            ("3.8.7", "MANUAL_REVIEW_REQUIRED",
             "Removable media controls cannot be assessed via network scan. Manual review of endpoint policies required.",
             "Implement technical controls to manage use of removable media on system components.",
             "medium"),
            ("3.8.8", "MANUAL_REVIEW_REQUIRED",
             "Unidentified portable storage device prohibition cannot be assessed via network scan. Manual review of policies required.",
             "Prohibit use of portable storage devices without identifiable owners.",
             "medium"),
            ("3.8.9", "MANUAL_REVIEW_REQUIRED",
             "Backup CUI confidentiality cannot be assessed via network scan. Manual review of backup storage and protection procedures required.",
             "Protect confidentiality of backup CUI at storage locations.",
             "high"),
            ("3.9.1", "MANUAL_REVIEW_REQUIRED",
             "Personnel screening processes cannot be assessed via network scan. Manual review of HR screening procedures required.",
             "Implement personnel screening procedures for individuals requiring access to CUI systems.",
             "high"),
            ("3.9.2", "MANUAL_REVIEW_REQUIRED",
             "Personnel termination and transfer protections cannot be assessed via network scan. Manual review of offboarding procedures required.",
             "Implement procedures to protect CUI systems during personnel terminations and transfers.",
             "high"),
            ("3.10.1", "MANUAL_REVIEW_REQUIRED",
             "Physical access authorizations cannot be assessed via network scan. Manual review of physical access control systems required.",
             "Implement physical access controls to limit access to authorized individuals.",
             "high"),
            ("3.10.2", "MANUAL_REVIEW_REQUIRED",
             "Physical facility monitoring cannot be assessed via network scan. Manual review of physical security controls required.",
             "Protect and monitor the physical facility and support infrastructure.",
             "high"),
            ("3.10.3", "MANUAL_REVIEW_REQUIRED",
             "Visitor escort procedures cannot be assessed via network scan. Manual review of visitor management policies required.",
             "Implement and enforce visitor escort and monitoring procedures.",
             "medium"),
            ("3.10.4", "MANUAL_REVIEW_REQUIRED",
             "Physical access logs cannot be assessed via network scan. Manual review of physical access logging systems required.",
             "Maintain and review audit logs of physical access to facilities.",
             "medium"),
            ("3.10.5", "MANUAL_REVIEW_REQUIRED",
             "Physical access device management cannot be assessed via network scan. Manual review of key/badge management procedures required.",
             "Implement controls to manage physical access devices (keys, badges, etc.).",
             "medium"),
            ("3.10.6", "MANUAL_REVIEW_REQUIRED",
             "Alternate work site safeguards cannot be assessed via network scan. Manual review of telework policies required.",
             "Enforce safeguarding measures for CUI at alternate work sites.",
             "medium"),
            ("3.11.1", "MANUAL_REVIEW_REQUIRED",
             "Risk assessment processes cannot be assessed via network scan. Manual review of risk assessment documentation required.",
             "Conduct and document periodic organizational risk assessments.",
             "high"),
            ("3.11.2", "NOT_ASSESSED",
             "Vulnerability scanning results require dedicated vulnerability scanner output for full assessment. Network scan provides partial data only.",
             "Implement regular vulnerability scanning using an approved vulnerability scanner.",
             "high"),
            ("3.11.3", "MANUAL_REVIEW_REQUIRED",
             "Vulnerability remediation tracking cannot be assessed via network scan. Manual review of patch management records required.",
             "Implement vulnerability remediation processes aligned with risk assessment findings.",
             "high"),
            ("3.12.1", "MANUAL_REVIEW_REQUIRED",
             "Security assessment processes cannot be assessed via network scan. Manual review of assessment documentation required.",
             "Conduct periodic security control assessments to verify control effectiveness.",
             "high"),
            ("3.12.2", "MANUAL_REVIEW_REQUIRED",
             "Plans of action and milestones cannot be assessed via network scan. Manual review of POA&M documentation required.",
             "Develop and maintain plans of action and milestones for identified deficiencies.",
             "high"),
            ("3.12.3", "MANUAL_REVIEW_REQUIRED",
             "Ongoing security control monitoring cannot be assessed via network scan. Manual review of continuous monitoring program required.",
             "Implement continuous monitoring of security controls.",
             "high"),
            ("3.12.4", "MANUAL_REVIEW_REQUIRED",
             "System security plan cannot be assessed via network scan. Manual review of SSP documentation required.",
             "Develop, document, and maintain a system security plan (SSP).",
             "high"),
            ("3.13.2", "MANUAL_REVIEW_REQUIRED",
             "Architectural security design cannot be assessed via network scan. Manual review of system architecture documentation required.",
             "Apply security engineering principles in system design and development.",
             "medium"),
            ("3.13.3", "MANUAL_REVIEW_REQUIRED",
             "Security function isolation cannot be assessed via network scan. Manual review of system design required.",
             "Separate user functionality from system management functionality.",
             "medium"),
            ("3.13.4", "MANUAL_REVIEW_REQUIRED",
             "Information transfer via shared resources cannot be assessed via network scan. Manual review required.",
             "Implement controls to prevent unauthorized information transfer via shared system resources.",
             "medium"),
            ("3.13.5", "MANUAL_REVIEW_REQUIRED",
             "DMZ and subnet segmentation cannot be fully assessed via network scan. Manual review of network architecture required.",
             "Implement DMZ or logically separated subnetworks for publicly accessible components.",
             "high"),
            ("3.13.6", "MANUAL_REVIEW_REQUIRED",
             "Default-deny network policy cannot be fully assessed via network scan. Manual review of firewall rules required.",
             "Implement deny-all, permit-by-exception network communication policies.",
             "high"),
            ("3.13.7", "MANUAL_REVIEW_REQUIRED",
             "Split tunneling controls cannot be assessed via network scan. Manual review of VPN configuration required.",
             "Configure VPN and remote access systems to prevent split tunneling.",
             "high"),
            ("3.13.8", "MANUAL_REVIEW_REQUIRED",
             "Cryptographic protection of CUI in transmission cannot be fully assessed via network scan. Manual review of encryption configuration required.",
             "Implement approved cryptographic mechanisms for all CUI transmission.",
             "high"),
            ("3.13.9", "MANUAL_REVIEW_REQUIRED",
             "Network session termination settings cannot be fully assessed via network scan. Manual review of firewall and session configuration required.",
             "Configure network devices to terminate inactive connections after a defined period.",
             "medium"),
            ("3.13.10", "MANUAL_REVIEW_REQUIRED",
             "Cryptographic key management cannot be assessed via network scan. Manual review of PKI and key management procedures required.",
             "Establish and maintain cryptographic key management procedures.",
             "high"),
            ("3.13.11", "MANUAL_REVIEW_REQUIRED",
             "FIPS-validated cryptography usage cannot be assessed via network scan. Manual review of cryptographic implementations required.",
             "Ensure all cryptographic implementations use FIPS-validated modules.",
             "high"),
            ("3.13.12", "MANUAL_REVIEW_REQUIRED",
             "Collaborative computing device controls cannot be assessed via network scan. Manual review of policies and device configuration required.",
             "Prohibit remote activation of collaborative computing devices and provide usage indicators.",
             "medium"),
            ("3.13.13", "MANUAL_REVIEW_REQUIRED",
             "Mobile code controls cannot be assessed via network scan. Manual review of browser and application policies required.",
             "Implement controls to manage and monitor mobile code.",
             "medium"),
            ("3.13.14", "MANUAL_REVIEW_REQUIRED",
             "VoIP controls cannot be assessed via network scan. Manual review of VoIP system configuration required.",
             "Implement controls to manage and monitor VoIP technologies.",
             "medium"),
            ("3.13.15", "MANUAL_REVIEW_REQUIRED",
             "Communications session authenticity protection cannot be fully assessed via network scan. Manual review required.",
             "Implement mechanisms to protect authenticity of communications sessions.",
             "high"),
            ("3.13.16", "MANUAL_REVIEW_REQUIRED",
             "CUI confidentiality at rest cannot be assessed via network scan. Manual review of encryption-at-rest configuration required.",
             "Implement encryption for CUI stored at rest.",
             "high"),
            ("3.14.1", "MANUAL_REVIEW_REQUIRED",
             "Flaw remediation processes cannot be fully assessed via network scan. Manual review of patch management procedures required.",
             "Implement timely identification and remediation of system flaws.",
             "high"),
            ("3.14.2", "MANUAL_REVIEW_REQUIRED",
             "Malicious code protection cannot be fully assessed via network scan. Manual review of anti-malware configuration required.",
             "Deploy and maintain malicious code protection at all designated system locations.",
             "high"),
            ("3.14.3", "MANUAL_REVIEW_REQUIRED",
             "Security alert monitoring cannot be assessed via network scan. Manual review of threat intelligence and alerting processes required.",
             "Implement processes to monitor and respond to security alerts and advisories.",
             "high"),
            ("3.14.4", "MANUAL_REVIEW_REQUIRED",
             "Anti-malware update currency cannot be assessed via network scan. Manual review of update configuration required.",
             "Configure anti-malware tools to update automatically when new releases are available.",
             "high"),
            ("3.14.5", "MANUAL_REVIEW_REQUIRED",
             "Periodic and real-time scanning cannot be assessed via network scan. Manual review of anti-malware scan configuration required.",
             "Configure periodic system scans and real-time scanning of files from external sources.",
             "high"),
            ("3.14.6", "MANUAL_REVIEW_REQUIRED",
             "Inbound and outbound traffic monitoring cannot be fully assessed via network scan. Manual review of IDS/IPS configuration required.",
             "Implement monitoring of inbound and outbound communications for attacks and indicators.",
             "high"),
            ("3.14.7", "MANUAL_REVIEW_REQUIRED",
             "Unauthorized system use identification cannot be assessed via network scan. Manual review of monitoring and detection capabilities required.",
             "Implement capabilities to identify unauthorized use of organizational systems.",
             "high"),
        ]

        for control_id, status, finding, remediation, severity in manual_review_controls:
            if control_id not in auto_assessed_ids:
                results.append(ComplianceResult(
                    control_id=control_id,
                    control_name=self.controls[control_id]["name"],
                    control_text=self.controls[control_id]["text"],
                    status=status,
                    finding=finding,
                    remediation=remediation,
                    severity=severity,
                    evidence=["Manual review required - cannot be automatically assessed via network scan"]
                ))
                self._vresult(results[-1])

        # Catch any Rev 2 controls not covered by the hint list above
        hints_emitted = {entry[0] for entry in manual_review_controls}
        covered = auto_assessed_ids | hints_emitted
        for control_id, ctrl in self.controls.items():
            if control_id not in covered:
                results.append(ComplianceResult(
                    control_id=control_id,
                    control_name=ctrl["name"],
                    control_text=ctrl["text"],
                    status="MANUAL_REVIEW_REQUIRED",
                    finding=f"{ctrl['name']} cannot be automatically assessed via network scan. Manual review required.",
                    remediation=f"Review and implement controls for {ctrl['name']} per NIST SP 800-171 Rev 2.",
                    severity="medium",
                    evidence=["Manual review required - cannot be automatically assessed via network scan"]
                ))
                self._vresult(results[-1])

        if _verbose_mode:
            nc = sum(1 for r in results if r.status == 'non_compliant')
            manual = sum(1 for r in results if r.status == 'MANUAL_REVIEW_REQUIRED')
            passed = sum(1 for r in results if r.status == 'compliant')
            _vprint()
            _vprint(f"  Assessment summary: {len(results)} controls  |  "
                    f"Pass: {passed}  Fail: {nc}  Manual: {manual}")

        return results

    def _vresult(self, r: 'ComplianceResult') -> None:
        """Print a single compliance result row in verbose mode."""
        if not _verbose_mode:
            return
        status = _STATUS_LABEL.get(r.status, r.status)
        sev    = _SEV_LABEL.get(r.severity, r.severity)
        _vprint(f"{r.control_id:<14}{r.control_name:<38}{status:<18}{sev}")
        if r.status == 'non_compliant':
            _vprint(f"  {'':14}Finding: {r.finding[:72]}")

    def _check_weak_services(self, system: SystemInfo) -> List[str]:
        """Check for potentially weak or insecure services"""
        weak_services = []
        risky_ports = {20: "FTP", 21: "FTP", 22: "SSH", 23: "Telnet", 25: "SMTP", 53: "DNS", 135: "RPC", 139: "NetBIOS or SMB" , 445: "SMB", 1433: "MSSQL", 1434: "MSSQL", 3306: "MySQL", 3389: "RDP", 5900: "VNC", 5901: "VNC", 5902: "VNC"}

        for port in system.open_ports:
            if port in risky_ports:
                service_name = system.services.get(port, risky_ports[port])
                weak_services.append(f"{service_name} (port {port})")

        return weak_services

    def _check_unnecessary_ports(self, system: SystemInfo) -> List[int]:
        """Check for potentially unnecessary open ports"""
        essential_ports = {22, 80, 443}
        return [port for port in system.open_ports if port not in essential_ports]

    def _check_external_services(self, system: SystemInfo) -> List[str]:
        """Check for services that might be externally accessible"""
        external_services = []
        external_ports = {21, 22, 23, 80, 443, 993, 995}

        for port in system.open_ports:
            if port in external_ports:
                service_name = system.services.get(port, f"Port {port}")
                external_services.append(service_name)

        return external_services


class ComplianceAssessorRev3:
    """NIST SP 800-171 Rev 3 compliance assessor (132 controls, 17 families)"""

    def __init__(self):
        self.controls = NIST80171ControlsRev3.CONTROLS
        self._auto = {
            "access_enforcement": "03.01.02",
            "nonessential_ports":  "03.04.07",
            "boundary_protection": "03.13.01",
        }

    def assess_system(self, system: SystemInfo) -> List[ComplianceResult]:
        """Assess a single system against NIST SP 800-171 Rev 3 controls"""
        results = []
        auto_assessed_ids = set(self._auto.values())

        logger.info(f"[ASSESS] [Rev 3] Assessing {system.hostname} ({system.ip_address})")
        _vsubsection(f"COMPLIANCE ASSESSMENT  |  {system.hostname}  ({system.ip_address})  [Rev 3]")
        _vprint(f"{'Control ID':<14}{'Control Name':<38}{'Status':<18}{'Severity'}")
        _vprint("-" * 78)

        # --- Auto-assessed: Access Enforcement (weak/insecure services) ---
        ctrl_ae = self._auto["access_enforcement"]
        weak_services = self._check_weak_services(system)
        if weak_services:
            results.append(ComplianceResult(
                control_id=ctrl_ae,
                control_name=self.controls[ctrl_ae]["name"],
                control_text=self.controls[ctrl_ae]["text"],
                status="non_compliant",
                finding=f"Potentially insecure services detected: {', '.join(weak_services)}",
                remediation="Disable unnecessary services, implement strong authentication, and restrict access",
                severity="high",
                evidence=[f"Open ports: {system.open_ports}", f"Services: {system.services}"]
            ))
            self._vresult(results[-1])
        else:
            results.append(ComplianceResult(
                control_id=ctrl_ae,
                control_name=self.controls[ctrl_ae]["name"],
                control_text=self.controls[ctrl_ae]["text"],
                status="compliant",
                finding="No obviously insecure services detected",
                remediation="Continue monitoring for unauthorized services",
                severity="low",
                evidence=[f"Services reviewed: {list(system.services.values())}"]
            ))
            self._vresult(results[-1])

        # --- Auto-assessed: Nonessential Functionality (unnecessary open ports) ---
        ctrl_np = self._auto["nonessential_ports"]
        unnecessary_ports = self._check_unnecessary_ports(system)
        if unnecessary_ports:
            results.append(ComplianceResult(
                control_id=ctrl_np,
                control_name=self.controls[ctrl_np]["name"],
                control_text=self.controls[ctrl_np]["text"],
                status="non_compliant",
                finding=f"Potentially unnecessary ports open: {unnecessary_ports}",
                remediation="Review and close unnecessary ports, disable unused services",
                severity="medium",
                evidence=[f"Open ports: {system.open_ports}"]
            ))
            self._vresult(results[-1])

        # --- Auto-assessed: Boundary Protection (external-facing services) ---
        ctrl_bp = self._auto["boundary_protection"]
        external_services = self._check_external_services(system)
        if external_services:
            results.append(ComplianceResult(
                control_id=ctrl_bp,
                control_name=self.controls[ctrl_bp]["name"],
                control_text=self.controls[ctrl_bp]["text"],
                status="non_compliant",
                finding=f"External-facing services detected: {external_services}",
                remediation="Implement firewall rules, access controls, and monitoring for external-facing services",
                severity="high",
                evidence=[f"External services: {external_services}"]
            ))
            self._vresult(results[-1])

        # All remaining Rev 3 controls require manual review
        for control_id, ctrl in self.controls.items():
            if control_id not in auto_assessed_ids:
                results.append(ComplianceResult(
                    control_id=control_id,
                    control_name=ctrl["name"],
                    control_text=ctrl["text"],
                    status="MANUAL_REVIEW_REQUIRED",
                    finding=f"{ctrl['name']} cannot be automatically assessed via network scan. Manual review required.",
                    remediation=f"Review and implement controls for {ctrl['name']} per NIST SP 800-171 Rev 3.",
                    severity="medium",
                    evidence=["Manual review required - cannot be automatically assessed via network scan"]
                ))
                self._vresult(results[-1])

        if _verbose_mode:
            nc = sum(1 for r in results if r.status == 'non_compliant')
            manual = sum(1 for r in results if r.status == 'MANUAL_REVIEW_REQUIRED')
            passed = sum(1 for r in results if r.status == 'compliant')
            _vprint()
            _vprint(f"  Assessment summary: {len(results)} controls  |  "
                    f"Pass: {passed}  Fail: {nc}  Manual: {manual}")

        return results

    def _vresult(self, r: 'ComplianceResult') -> None:
        """Print a single compliance result row in verbose mode."""
        if not _verbose_mode:
            return
        status = _STATUS_LABEL.get(r.status, r.status)
        sev    = _SEV_LABEL.get(r.severity, r.severity)
        _vprint(f"{r.control_id:<14}{r.control_name:<38}{status:<18}{sev}")
        if r.status == 'non_compliant':
            _vprint(f"  {'':14}Finding: {r.finding[:72]}")

    def _check_weak_services(self, system: SystemInfo) -> List[str]:
        """Check for potentially weak or insecure services"""
        weak_services = []
        risky_ports = {21: "FTP", 23: "Telnet", 135: "RPC", 139: "NetBIOS", 445: "SMB"}
        for port in system.open_ports:
            if port in risky_ports:
                service_name = system.services.get(port, risky_ports[port])
                weak_services.append(f"{service_name} (port {port})")
        return weak_services

    def _check_unnecessary_ports(self, system: SystemInfo) -> List[int]:
        """Check for potentially unnecessary open ports"""
        essential_ports = {22, 80, 443}
        return [port for port in system.open_ports if port not in essential_ports]

    def _check_external_services(self, system: SystemInfo) -> List[str]:
        """Check for services that might be externally accessible"""
        external_services = []
        external_ports = {21, 22, 23, 80, 443, 993, 995}
        for port in system.open_ports:
            if port in external_ports:
                service_name = system.services.get(port, f"Port {port}")
                external_services.append(service_name)
        return external_services


class SPRSCalculator:
    """Calculate SPRS (Supplier Performance Risk System) score"""

    def __init__(self, controls: Dict = None):
        self.controls = controls if controls is not None else NIST80171Controls.CONTROLS

    def calculate_sprs_score(self, results: List[ComplianceResult]) -> Dict:
        """Calculate SPRS score based on compliance results"""
        logger.info("[SPRS] Calculating SPRS compliance score...")

        total_controls = len(self.get_all_control_ids(self.controls))

        compliant = len([r for r in results if r.status == 'compliant'])
        non_compliant = len([r for r in results if r.status == 'non_compliant'])
        not_applicable = len([r for r in results if r.status == 'not_applicable'])
        not_assessed = len([r for r in results if r.status == 'not_assessed'])

        applicable_controls = total_controls - not_applicable
        if applicable_controls > 0:
            compliance_percentage = (compliant / applicable_controls) * 100
        else:
            compliance_percentage = 100

        base_score = 110

        high_severity_deduction = len([r for r in results if r.status == 'non_compliant' and r.severity == 'high']) * 15
        medium_severity_deduction = len([r for r in results if r.status == 'non_compliant' and r.severity == 'medium']) * 10
        low_severity_deduction = len([r for r in results if r.status == 'non_compliant' and r.severity == 'low']) * 5

        total_deduction = high_severity_deduction + medium_severity_deduction + low_severity_deduction
        sprs_score = max(0, base_score - total_deduction)

        result_dict = {
            'sprs_score': sprs_score,
            'max_score': base_score,
            'compliance_percentage': round(compliance_percentage, 2),
            'total_controls': total_controls,
            'compliant': compliant,
            'non_compliant': non_compliant,
            'not_applicable': not_applicable,
            'not_assessed': not_assessed,
            'high_severity_findings': len([r for r in results if r.status == 'non_compliant' and r.severity == 'high']),
            'medium_severity_findings': len([r for r in results if r.status == 'non_compliant' and r.severity == 'medium']),
            'low_severity_findings': len([r for r in results if r.status == 'non_compliant' and r.severity == 'low'])
        }

        if _verbose_mode:
            _vsection("SPRS SCORE RESULTS")
            _vkv("SPRS Score",        f"{sprs_score} / {base_score}")
            _vkv("Compliance Rate",   f"{round(compliance_percentage, 2)}%")
            _vkv("Total Controls",    total_controls)
            _vkv("Compliant",         compliant)
            _vkv("Non-Compliant",     non_compliant)
            _vkv("Not Assessed",      not_assessed)
            _vkv("High Findings",     result_dict['high_severity_findings'])
            _vkv("Medium Findings",   result_dict['medium_severity_findings'])
            _vkv("Low Findings",      result_dict['low_severity_findings'])

        return result_dict

    def get_all_control_ids(self, controls: Dict = None) -> List[str]:
        """Get all NIST SP 800-171 control IDs for the active version"""
        if controls is not None:
            return list(controls.keys())
        return list(NIST80171Controls.CONTROLS.keys())

class ReportGenerator:
    """Generate compliance reports in multiple formats"""

    def __init__(self, output_dir: str = "reports", standard_label: str = "NIST SP 800-171 Rev 2"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)
        self.standard_label = standard_label

    def generate_all_reports(self, systems: List[SystemInfo], results: List[ComplianceResult], sprs_data: Dict, topology: NetworkTopology = None):
        """Generate reports in all formats"""
        logger.info("[REPORTS] Generating compliance reports in multiple formats...")
        _vsection("REPORT GENERATION")
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")

        html_file = f"nistify_compliance_report_{timestamp}.html"
        json_file = f"nistify_compliance_report_{timestamp}.json"
        txt_file  = f"nistify_compliance_report_{timestamp}.txt"
        xlsx_file = f"nistify_poam_{timestamp}.xlsx"
        nmap_file = f"nistify_scan_{timestamp}.nmap"

        _vprint(f"[>] Generating HTML report  ...")
        self.generate_html_report(systems, results, sprs_data, html_file, topology)
        _vprint(f"[+] {self.output_dir / html_file}")

        _vprint(f"[>] Generating JSON report  ...")
        self.generate_json_report(systems, results, sprs_data, json_file, topology)
        _vprint(f"[+] {self.output_dir / json_file}")

        _vprint(f"[>] Generating text report  ...")
        self.generate_text_report(systems, results, sprs_data, txt_file)
        _vprint(f"[+] {self.output_dir / txt_file}")

        _vprint(f"[>] Generating POA&M (Excel)...")
        self.generate_poam_xlsx(results, xlsx_file)
        _vprint(f"[+] {self.output_dir / xlsx_file}")

        _vprint(f"[>] Generating Nmap report  ...")
        self.generate_nmap_report(systems, nmap_file)
        _vprint(f"[+] {self.output_dir / nmap_file}")

        logger.info(f"[REPORTS] Reports generated in {self.output_dir}")

    def generate_nmap_report(self, systems: List[SystemInfo], filename: str):
        """Generate nmap-style .nmap text report for all scanned hosts"""
        logger.info("[NMAP] Generating nmap-format report...")
        scan_time = datetime.datetime.now()
        scan_time_str = scan_time.strftime("%a %b %d %H:%M:%S %Y")

        lines = []
        lines.append(f"# Nmap scan report generated by NISTify v2.0.0")
        lines.append(f"# {self.standard_label} Compliance Scan")
        lines.append(f"# Scan initiated {scan_time_str}")
        lines.append("")

        for system in systems:
            lines.append(f"Nmap scan report for {system.hostname} ({system.ip_address})")
            lines.append(f"Host is up.")

            if system.hop_count is not None:
                lines.append(f"Distance: {system.hop_count} hops")

            if system.mac_address:
                vendor_str = f" ({system.vendor})" if system.vendor else ""
                lines.append(f"MAC Address: {system.mac_address}{vendor_str}")

            if system.open_ports:
                lines.append("")
                lines.append("PORT      STATE  SERVICE")
                for port in sorted(system.open_ports):
                    service = system.services.get(port, "unknown")
                    port_proto = f"{port}/tcp"
                    lines.append(f"{port_proto:<10}open   {service}")
            else:
                lines.append("All scanned ports are closed or filtered.")

            if system.os_type and system.os_type != "Unknown":
                lines.append("")
                os_detail = system.os_type
                if system.os_version and system.os_version != "Unknown":
                    os_detail += f" {system.os_version}"
                lines.append(f"OS details: {os_detail}")

            lines.append("")

        total_hosts = len(systems)
        lines.append(f"# Nmap done: {total_hosts} IP address{'es' if total_hosts != 1 else ''} ({total_hosts} host{'s' if total_hosts != 1 else ''} up) scanned")
        lines.append(f"# End of report: {scan_time_str}")

        with open(self.output_dir / filename, 'w', encoding='utf-8') as f:
            f.write('\n'.join(lines) + '\n')

        logger.info(f"[NMAP] Nmap report generated: {filename}")

    def generate_html_report(self, systems: List[SystemInfo], results: List[ComplianceResult], sprs_data: Dict, filename: str, topology: NetworkTopology = None):
        """Generate HTML compliance report"""
        html_content = f"""
<!DOCTYPE html>
<html>
<head>
    <title>NISTify - {self.standard_label} Compliance Report</title>
    <style>
        body {{ font-family: Arial, sans-serif; margin: 20px; background-color: #f8f9fa; }}
        .header {{ background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 30px; border-radius: 10px; text-align: center; }}
        .summary {{ background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 25px; margin: 20px 0; border-radius: 10px; }}
        .sprs-score {{ font-size: 36px; font-weight: bold; text-align: center; background: rgba(255,255,255,0.2); padding: 20px; border-radius: 10px; margin: 15px 0; }}
        table {{ border-collapse: collapse; width: 100%; margin: 20px 0; background: white; border-radius: 10px; overflow: hidden; }}
        th, td {{ border: none; padding: 12px 15px; text-align: left; }}
        th {{ background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; font-weight: bold; }}
        tr:nth-child(even) {{ background-color: #f8f9fa; }}
        .compliant {{ background-color: #d4edda !important; }}
        .non-compliant {{ background-color: #f8d7da !important; }}
        .high-severity {{ color: #dc3545; font-weight: bold; }}
        .medium-severity {{ color: #fd7e14; font-weight: bold; }}
        .low-severity {{ color: #28a745; font-weight: bold; }}
    </style>
</head>
<body>
    <div class="header">
        <h1>NISTify - {self.standard_label}</h1>
        <p>{self.standard_label} Compliance Assessment Report</p>
        <p>Generated on: {datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")}</p>
    </div>

    <div class="summary">
        <h2>Executive Summary</h2>
        <div class="sprs-score">
            SPRS Score: {sprs_data['sprs_score']} / {sprs_data['max_score']}
            <div style="font-size: 18px; margin-top: 10px;">
                Compliance: {sprs_data['compliance_percentage']}%
            </div>
        </div>
        <p>Systems Assessed: {len(systems)}</p>
        <p>Total Findings: {len(results)}</p>
        <p>High Severity: {sprs_data['high_severity_findings']}, Medium: {sprs_data['medium_severity_findings']}, Low: {sprs_data['low_severity_findings']}</p>
    </div>

    <h2>Scanned Systems</h2>
    <table>
        <tr>
            <th>Hostname</th>
            <th>IP Address</th>
            <th>OS Type</th>
            <th>Open Ports</th>
            <th>Last Scanned</th>
        </tr>
"""

        for system in systems:
            html_content += f"""
        <tr>
            <td><strong>{system.hostname}</strong></td>
            <td>{system.ip_address}</td>
            <td>{system.os_type}</td>
            <td>{', '.join(map(str, system.open_ports))}</td>
            <td>{system.last_scanned}</td>
        </tr>
"""

        html_content += """
    </table>

    <h2>Compliance Findings</h2>
    <table>
        <tr>
            <th>Control ID</th>
            <th>Control Name</th>
            <th>Status</th>
            <th>Severity</th>
            <th>Finding</th>
            <th>Remediation</th>
        </tr>
"""

        for result in results:
            status_class = result.status.replace('_', '-')
            severity_class = f"{result.severity}-severity"
            html_content += f"""
        <tr class="{status_class}">
            <td><strong>{result.control_id}</strong></td>
            <td>{result.control_name}</td>
            <td>{result.status.replace('_', ' ').title()}</td>
            <td class="{severity_class}">{result.severity.title()}</td>
            <td>{result.finding}</td>
            <td>{result.remediation}</td>
        </tr>
"""

        html_content += """
    </table>

    <div style="text-align: center; margin-top: 40px; padding: 20px; background: #f8f9fa; border-radius: 10px;">
        <p style="color: #666; margin: 0;">
            Generated by NISTify v2.0.0 |
            {self.standard_label} Compliance Assessment Tool
        </p>
    </div>
</body>
</html>
        """

        with open(self.output_dir / filename, 'w', encoding='utf-8') as f:
            f.write(html_content)

    def generate_json_report(self, systems: List[SystemInfo], results: List[ComplianceResult], sprs_data: Dict, filename: str, topology: NetworkTopology = None):
        """Generate JSON compliance report"""
        report_data = {
            "metadata": {
                "generated_on": datetime.datetime.now().isoformat(),
                "standard": self.standard_label,
                "tool": "NISTify",
                "version": "2.0.0"
            },
            "sprs_score": sprs_data,
            "scanned_systems": [
                {**asdict(system), 'services': {str(k): v for k, v in system.services.items()}}
                for system in systems
            ],
            "compliance_results": [asdict(result) for result in results]
        }

        if topology:
            report_data["network_topology"] = {
                "nodes": topology.nodes,
                "edges": topology.edges,
                "subnets": topology.subnets,
                "gateways": topology.gateways,
                "diagram_path": topology.network_diagram_path
            }

        with open(self.output_dir / filename, 'w', encoding='utf-8') as f:
            json.dump(report_data, f, indent=2, ensure_ascii=False)

    def generate_text_report(self, systems: List[SystemInfo], results: List[ComplianceResult], sprs_data: Dict, filename: str):
        """Generate text compliance report"""
        with open(self.output_dir / filename, 'w', encoding='utf-8') as f:
            f.write("╔══════════════════════════════════════════════════════════════════════════════════════════════╗\n")
            f.write(f"║                     NISTify {self.standard_label} COMPLIANCE REPORT                    ║\n")
            f.write("╚══════════════════════════════════════════════════════════════════════════════════════════════╝\n\n")

            f.write(f"Generated on: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"Standard: {self.standard_label}\n")
            f.write("Tool: NISTify v2.0.0\n\n")

            f.write("EXECUTIVE SUMMARY\n")
            f.write("=" * 50 + "\n")
            f.write(f"SPRS Score: {sprs_data['sprs_score']} / {sprs_data['max_score']}\n")
            f.write(f"Compliance Percentage: {sprs_data['compliance_percentage']}%\n")
            f.write(f"Systems Assessed: {len(systems)}\n")
            f.write(f"Total Findings: {len(results)}\n")
            f.write(f"High Severity: {sprs_data['high_severity_findings']}\n")
            f.write(f"Medium Severity: {sprs_data['medium_severity_findings']}\n")
            f.write(f"Low Severity: {sprs_data['low_severity_findings']}\n\n")

            f.write("SCANNED SYSTEMS\n")
            f.write("=" * 30 + "\n")
            for system in systems:
                f.write(f"Hostname: {system.hostname}\n")
                f.write(f"IP Address: {system.ip_address}\n")
                f.write(f"OS Type: {system.os_type}\n")
                f.write(f"Open Ports: {', '.join(map(str, system.open_ports))}\n")
                f.write(f"Last Scanned: {system.last_scanned}\n\n")

            f.write("COMPLIANCE FINDINGS\n")
            f.write("=" * 35 + "\n")
            for result in results:
                f.write(f"Control ID: {result.control_id}\n")
                f.write(f"Control Name: {result.control_name}\n")
                f.write(f"Status: {result.status.replace('_', ' ').title()}\n")
                f.write(f"Severity: {result.severity.title()}\n")
                f.write(f"Finding: {result.finding}\n")
                f.write(f"Remediation: {result.remediation}\n")
                if result.evidence:
                    f.write(f"Evidence: {'; '.join(result.evidence)}\n")
                f.write("\n" + "-" * 80 + "\n\n")

            f.write("\n" + "="*80 + "\n")
            f.write("Generated by NISTify v2.0.0\n")
            f.write(f"Comprehensive {self.standard_label} Compliance Assessment Tool\n")
            f.write("="*80 + "\n")

    def generate_poam_xlsx(self, results: List[ComplianceResult], filename: str):
        """Generate Plan of Action and Milestones (POA&M) Excel document"""
        logger.info("[POAM] Generating POA&M Excel document...")

        wb = Workbook()
        ws = wb.active
        ws.title = "NISTify POA&M"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
        high_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
        medium_fill = PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid")
        low_fill = PatternFill(start_color="6BCF7F", end_color="6BCF7F", fill_type="solid")

        headers = [
            "Control Number", "Control Name", "Control Text", "Status", "Severity",
            "Deficiency Identified", "Remediation Steps", "Target Date",
            "Responsible Party", "Status Notes", "Evidence"
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        row = 2
        for result in results:
            if result.status == 'non_compliant':
                ws.cell(row, 1, result.control_id)
                ws.cell(row, 2, result.control_name)
                ws.cell(row, 3, result.control_text)
                ws.cell(row, 4, result.status.replace('_', ' ').title())

                severity_cell = ws.cell(row, 5, result.severity.title())
                if result.severity == 'high':
                    severity_cell.fill = high_fill
                elif result.severity == 'medium':
                    severity_cell.fill = medium_fill
                else:
                    severity_cell.fill = low_fill

                ws.cell(row, 6, result.finding)
                ws.cell(row, 7, result.remediation)

                target_date = datetime.datetime.now()
                if result.severity == 'high':
                    target_date += datetime.timedelta(days=30)
                elif result.severity == 'medium':
                    target_date += datetime.timedelta(days=90)
                else:
                    target_date += datetime.timedelta(days=180)

                ws.cell(row, 8, target_date.strftime("%Y-%m-%d"))
                ws.cell(row, 9, "IT Security Team")
                ws.cell(row, 10, "Open")
                ws.cell(row, 11, '; '.join(result.evidence) if result.evidence else "")

                row += 1

        column_widths = [15, 30, 50, 15, 10, 40, 50, 12, 20, 15, 30]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(1, col).column_letter].width = width

        ws.auto_filter.ref = f"A1:{ws.cell(1, len(headers)).coordinate}"

        wb.save(self.output_dir / filename)
        logger.info(f"[POAM] POA&M Excel document generated: {filename}")

class ComplianceScanner:
    """Main NISTify compliance scanner orchestrator"""

    def __init__(self, output_dir: str = "reports", assessor=None, version: str = "Rev 2", standard_label: str = "NIST SP 800-171 Rev 2"):
        self.version = version
        self.scanner = NetworkScanner()
        self.assessor = assessor
        self.sprs_calculator = SPRSCalculator(controls=assessor.controls)
        self.report_generator = ReportGenerator(output_dir, standard_label=standard_label)

    def scan_and_assess(self, network_ranges: List[str], generate_topology: bool = True) -> Tuple[List[SystemInfo], List[ComplianceResult], Dict, NetworkTopology]:
        """Perform complete scan and assessment"""
        logger.info(f"[START] Starting NIST SP 800-171 {self.version} compliance assessment...")

        all_systems = []
        all_results = []
        topology = None

        for network_range in network_ranges:
            systems = self.scanner.scan_network(network_range)
            all_systems.extend(systems)

        if generate_topology and all_systems:
            _vsection(f"NETWORK TOPOLOGY ANALYSIS  |  {len(all_systems)} host(s)")
            topology = self.scanner.discover_topology(all_systems)
            diagram_path = str(Path(self.report_generator.output_dir) / "nistify_network_topology.png")
            self.scanner.create_network_diagram(topology, diagram_path)

        logger.info(f"[ASSESS] Assessing {len(all_systems)} systems for NIST SP 800-171 compliance")
        _vsection(f"COMPLIANCE ASSESSMENT  |  {self.version}  |  {len(all_systems)} system(s)")
        for system in all_systems:
            results = self.assessor.assess_system(system)
            all_results.extend(results)

        # Deduplicate: one finding per control ID, keeping the most severe status
        # so that a non-compliant result on any host surfaces in the final report.
        _severity_rank = {"non_compliant": 3, "MANUAL_REVIEW_REQUIRED": 2,
                          "not_assessed": 1, "not_applicable": 0, "compliant": 0}
        deduped: Dict[str, ComplianceResult] = {}
        for r in all_results:
            existing = deduped.get(r.control_id)
            if existing is None or _severity_rank.get(r.status, 0) > _severity_rank.get(existing.status, 0):
                deduped[r.control_id] = r
        all_results = list(deduped.values())

        sprs_data = self.sprs_calculator.calculate_sprs_score(all_results)

        return all_systems, all_results, sprs_data, topology

    def generate_reports(self, systems: List[SystemInfo], results: List[ComplianceResult], sprs_data: Dict, topology: NetworkTopology = None):
        """Generate all compliance reports"""
        self.report_generator.generate_all_reports(systems, results, sprs_data, topology)

def main():
    """Main entry point"""
    print_banner()

    version_label, active_controls, standard_label = select_nist_version()
    print(f"\n  Using {standard_label} ({len(active_controls)} controls)\n")

    parser = argparse.ArgumentParser(
        description=f"NISTify 800-171 - {standard_label} Compliance Scanner",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python nistify800-171r2.py 192.168.1.0/24
  python nistify800-171r2.py 10.0.0.0/8 192.168.0.0/16 --verbose
  python nistify800-171r2.py 172.16.0.0/12 --no-topology
        """
    )

    parser.add_argument("networks", nargs="+",
                       help="Network ranges to scan (e.g., 192.168.1.0/24)")
    parser.add_argument("--output-dir", default="nistify_reports",
                       help="Output directory for reports (default: nistify_reports)")
    parser.add_argument("--no-topology", action="store_true",
                       help="Skip network topology generation for faster scanning")
    parser.add_argument("--verbose", "-v", action="store_true", help="Enable verbose logging")

    args = parser.parse_args()

    if args.verbose:
        global _verbose_mode
        _verbose_mode = True
        logging.getLogger().setLevel(logging.DEBUG)
        logger.info("[CONFIG] Verbose logging enabled")
        _vsection("NISTIFY 800-171  —  SCAN CONFIGURATION")
        _vkv("Standard",    standard_label)
        _vkv("Controls",    len(active_controls))
        _vkv("Networks",    ', '.join(args.networks))
        _vkv("Output Dir",  args.output_dir)
        _vkv("Topology",    "Disabled" if args.no_topology else "Enabled")
        _vkv("Verbose",     "Enabled")

    valid_networks = []
    logger.info("[CONFIG] Validating network ranges...")
    for network in args.networks:
        try:
            ipaddress.ip_network(network, strict=False)
            valid_networks.append(network)
            logger.info(f"[CONFIG] Valid network range: {network}")
        except ValueError:
            logger.error(f"[ERROR] Invalid network range: {network}")
            continue

    if not valid_networks:
        logger.error("[ERROR] No valid network ranges provided")
        sys.exit(1)

    if version_label == "Rev 2":
        assessor = ComplianceAssessorRev2()
    else:
        assessor = ComplianceAssessorRev3()

    scanner = ComplianceScanner(args.output_dir, assessor=assessor, version=version_label, standard_label=standard_label)

    try:
        start_time = datetime.datetime.now()

        systems, results, sprs_data, topology = scanner.scan_and_assess(
            valid_networks, not args.no_topology
        )

        end_time = datetime.datetime.now()
        duration = end_time - start_time

        logger.info(f"[COMPLETE] Assessment complete in {duration.total_seconds():.1f} seconds")
        logger.info(f"[RESULTS] Found {len(systems)} systems with {len(results)} findings")
        logger.info(f"[SPRS] SPRS Score: {sprs_data['sprs_score']} / {sprs_data['max_score']}")

        if topology:
            logger.info(f"[TOPOLOGY] Network topology: {len(topology.nodes)} nodes, {len(topology.edges)} connections")
            logger.info(f"[TOPOLOGY] Discovered subnets: {', '.join(topology.subnets)}")

        scanner.generate_reports(systems, results, sprs_data, topology)

        print(f"\n{'='*90}")
        print(f"NISTify 800-171 {version_label} COMPLIANCE ASSESSMENT COMPLETE")
        print(f"{'='*90}")
        print(f"Assessment Duration: {duration.total_seconds():.1f} seconds")
        print(f"Systems Scanned: {len(systems)}")
        print(f"Compliance Findings: {len(results)}")
        print(f"SPRS Score: {sprs_data['sprs_score']} / {sprs_data['max_score']}")
        print(f"Compliance Rate: {sprs_data['compliance_percentage']}%")
        print(f"High Severity Issues: {sprs_data['high_severity_findings']}")
        print(f"Medium Severity Issues: {sprs_data['medium_severity_findings']}")
        print(f"Low Severity Issues: {sprs_data['low_severity_findings']}")

        if topology:
            print("\nNetwork Topology Analysis:")
            print(f"   Total Nodes: {len(topology.nodes)}")
            print(f"   Network Connections: {len(topology.edges)}")
            print(f"   Subnets Discovered: {len(topology.subnets)}")
            print(f"   Gateways Identified: {len(topology.gateways)}")

        print(f"\nReports generated in: {args.output_dir}")
        print("   HTML Report: nistify_compliance_report_*.html")
        print("   JSON Report: nistify_compliance_report_*.json")
        print("   Text Report: nistify_compliance_report_*.txt")
        print("   POA&M Document: nistify_poam_*.xlsx")
        print("   Nmap Report: nistify_scan_*.nmap")

        if topology and topology.network_diagram_path:
            print("   Network Topology Diagram: nistify_network_topology.png")

        print(f"\nThank you for using NISTify 800-171 {version_label}!")
        print("   For support and updates: https://github.com/nightstalker117/nistify-800-171")

    except KeyboardInterrupt:
        logger.info("\n[INTERRUPTED] Assessment interrupted by user")
        sys.exit(0)
    except Exception as e:
        logger.error(f"[ERROR] Assessment failed: {e}")
        if args.verbose:
            import traceback
            logger.error(f"Full traceback: {traceback.format_exc()}")
        sys.exit(1)

if __name__ == "__main__":
    main()
